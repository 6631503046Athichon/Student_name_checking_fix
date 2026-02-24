"""
สคริปต์สร้างไฟล์ Excel ตัวอย่างสำหรับทดสอบการ Import
"""

import openpyxl
from openpyxl import Workbook
import os


def create_sample_students_excel():
    """สร้างไฟล์ Excel ตัวอย่างรายชื่อนักเรียน"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Header
    headers = [
        'รหัสนักเรียน', 'คำนำหน้า', 'ชื่อ', 'นามสกุล',
        'ห้อง', 'ปีการศึกษา', 'วันเกิด',
        'ชื่อผู้ปกครอง', 'เบอร์โทรผู้ปกครอง'
    ]
    ws.append(headers)

    # Sample data
    students = [
        ['67001', 'เด็กชาย', 'สมชาย', 'ใจดี', 'ป.1/1', '2567', '2016-01-15', 'นายสมศักดิ์ ใจดี', '0812345678'],
        ['67002', 'เด็กหญิง', 'สมหญิง', 'รักเรียน', 'ป.1/1', '2567', '2016-02-20', 'นางสมใจ รักเรียน', '0823456789'],
        ['67003', 'เด็กชาย', 'ธนา', 'มีปัญญา', 'ป.2/1', '2567', '2015-05-10', 'นายธนากร มีปัญญา', '0834567890'],
        ['67004', 'เด็กหญิง', 'วิภา', 'ฉลาด', 'ป.2/1', '2567', '2015-08-25', 'นางวิมล ฉลาด', '0845678901'],
        ['67005', 'เด็กชาย', 'ชัยวัฒน์', 'เก่งกาจ', 'ป.3/1', '2567', '2014-12-05', 'นายชัยยุทธ เก่งกาจ', '0856789012'],
    ]

    for student in students:
        ws.append(student)

    # บันทึกไฟล์
    file_path = os.path.join(os.path.dirname(__file__), 'students.xlsx')
    wb.save(file_path)
    print(f"Created {file_path}")


def create_sample_teachers_excel():
    """สร้างไฟล์ Excel ตัวอย่างรายชื่อครู"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"

    # Header
    headers = ['รหัสครู', 'คำนำหน้า', 'ชื่อ', 'นามสกุล', 'เบอร์โทร']
    ws.append(headers)

    # Sample data
    teachers = [
        ['T101', 'นาย', 'สมศักดิ์', 'สอนดี', '0891234567'],
        ['T102', 'นาง', 'สมหญิง', 'รักสอน', '0892345678'],
        ['T103', 'นางสาว', 'วิภา', 'ใจเย็น', '0893456789'],
    ]

    for teacher in teachers:
        ws.append(teacher)

    # บันทึกไฟล์
    file_path = os.path.join(os.path.dirname(__file__), 'teachers.xlsx')
    wb.save(file_path)
    print(f"Created {file_path}")


def create_invalid_excel():
    """สร้างไฟล์ Excel ที่มีรูปแบบผิดสำหรับทดสอบ error handling"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invalid"

    # Header ผิด
    headers = ['A', 'B', 'C']
    ws.append(headers)

    # ข้อมูลไม่ครบ
    ws.append(['67999', 'เด็กชาย'])  # ขาดคอลัมน์

    # บันทึกไฟล์
    file_path = os.path.join(os.path.dirname(__file__), 'students_invalid.xlsx')
    wb.save(file_path)
    print(f"Created {file_path}")


if __name__ == '__main__':
    create_sample_students_excel()
    create_sample_teachers_excel()
    create_invalid_excel()
