"""
modules/reports.py
โมดูลรายงาน - Design System v3.0
- Dashboard cards radius 12px
- Icon 48px
- ตัวเลข H1 (28px)
- Charts ใช้สีตาม design system
- Spacing 32px ระหว่าง sections
- ปุ่ม export ใช้ SECONDARY style (outlined)
- Toast notification แทน status bar
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from modules.icons import IconManager
from modules.pdf_utils import get_thai_font

# ==================== Design System v4.0 ====================
# Accent Colors (10%)
PRIMARY = "#3B82F6"
SUCCESS = "#10B981"
WARNING = "#F59E0B"
DANGER = "#EF4444"
NEUTRAL = "#64748B"

# Typography Colors
TEXT_H1 = "#0F172A"
TEXT_H2 = "#1E293B"
TEXT_H3 = "#334155"
TEXT_BODY = "#475569"
TEXT_CAPTION = "#6B7280"

# Surface / Background
SURFACE = "#FFFFFF"
BG_LIGHT = "#F5F7FA"

# Table
TABLE_HEADER_BG = "#F9FAFB"
TABLE_HOVER = "#EFF6FF"
TABLE_STRIPE = "#F9FAFB"
TABLE_BORDER = "#E5E7EB"

# Spacing
XS, S, M, L, XL, XXL = 4, 8, 16, 24, 32, 48

# Radius
RADIUS_BUTTON = 8
RADIUS_CARD = 12
RADIUS_MODAL = 16

# Input
INPUT_BORDER = "#D1D5DB"
INPUT_FOCUS = PRIMARY

# Report Card Color Mapping
CARD_COLORS = {
    "students": PRIMARY,
    "attendance": SUCCESS,
    "health": DANGER,
    "grades": WARNING,
    "schedule": "#7C3AED",
    "summary": NEUTRAL,
}


class ReportsModule:
    """โมดูลรายงาน - Design System v3.0"""

    def __init__(self, parent, db, update_status_callback):
        self.parent = parent
        self.db = db
        self.update_status = update_status_callback

        self.create_ui()

    def create_ui(self):
        """สร้าง UI ของโมดูล"""

        # Main scrollable frame - เลื่อนดูเนื้อหาได้เมื่อหน้าจอเล็ก
        main_frame = ctk.CTkScrollableFrame(
            self.parent,
            fg_color="transparent",
            scrollbar_button_color="#CBD5E1",
            scrollbar_button_hover_color=PRIMARY,
        )
        main_frame.pack(fill="both", expand=True, padx=L, pady=L)

        # ========== Header Section ==========
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, XL))

        ctk.CTkLabel(
            header_frame,
            text="📊 รายงานและส่งออกข้อมูล",
            font=ctk.CTkFont(family="Kanit", size=24, weight="600"),
            text_color=TEXT_H1,
        ).pack(side="left")

        ctk.CTkLabel(
            header_frame,
            text="เลือกประเภทรายงานเพื่อส่งออกข้อมูล",
            font=ctk.CTkFont(family="Kanit", size=14),
            text_color=TEXT_CAPTION,
        ).pack(side="left", padx=(M, 0))

        # ========== Dashboard Summary Cards ==========
        self._create_summary_cards(main_frame)

        # ========== Report Export Cards Section ==========
        section_label = ctk.CTkLabel(
            main_frame,
            text="ส่งออกรายงาน",
            font=ctk.CTkFont(family="TH Sarabun New", size=20, weight="bold"),
            text_color=TEXT_H2,
        )
        section_label.pack(anchor="w", pady=(XL, M))

        # Grid container
        grid_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        grid_frame.pack(fill="x")

        # Report cards definition
        reports = [
            {
                "key": "students",
                "title": "รายชื่อนักเรียน",
                "desc": "ข้อมูลนักเรียนทั้งหมด",
                "icon": "users",
                "excel": self.export_students_excel,
                "pdf": self.export_students_pdf,
            },
            {
                "key": "attendance",
                "title": "การเช็คชื่อ",
                "desc": "สถิติการเข้าเรียน",
                "icon": "clipboard-check",
                "excel": self.export_attendance_excel,
                "pdf": self.export_attendance_pdf,
            },
            {
                "key": "health",
                "title": "ข้อมูลสุขภาพ",
                "desc": "น้ำหนัก ส่วนสูง BMI",
                "icon": "heart-pulse",
                "excel": self.export_health_excel,
                "pdf": self.export_health_pdf,
            },
            {
                "key": "grades",
                "title": "เกรดและผลการเรียน",
                "desc": "ผลการเรียนทุกวิชา",
                "icon": "graduation-cap",
                "excel": self.export_grades_excel,
                "pdf": self.export_grades_pdf,
            },
            {
                "key": "schedule",
                "title": "ตารางเรียน/สอน",
                "desc": "ตารางสอนทุกห้อง",
                "icon": "calendar-days",
                "excel": self.export_schedule_excel,
                "pdf": self.export_schedule_pdf,
            },
            {
                "key": "summary",
                "title": "สรุปข้อมูลทั้งหมด",
                "desc": "รวมทุกรายงาน",
                "icon": "chart-pie",
                "excel": self.export_all_excel,
                "pdf": self.export_all_pdf,
            },
        ]

        # 3 columns grid
        grid_frame.columnconfigure(0, weight=1)
        grid_frame.columnconfigure(1, weight=1)
        grid_frame.columnconfigure(2, weight=1)

        for idx, report in enumerate(reports):
            row = idx // 3
            col = idx % 3
            self._create_report_card(grid_frame, report, row, col)

    def _create_summary_cards(self, parent):
        """สร้าง Dashboard Summary Cards ด้านบน"""

        cards_frame = ctk.CTkFrame(parent, fg_color="transparent")
        cards_frame.pack(fill="x", pady=(0, 0))

        cards_frame.columnconfigure(0, weight=1)
        cards_frame.columnconfigure(1, weight=1)
        cards_frame.columnconfigure(2, weight=1)

        # Fetch counts
        try:
            students = self.db.get_all_students()
            teachers = self.db.get_all_teachers()
            class_rooms = self.db.get_class_rooms()
            student_count = len(students)
            teacher_count = len(teachers)
            class_count = len(class_rooms)
        except Exception:
            student_count = 0
            teacher_count = 0
            class_count = 0

        summary_items = [
            {"label": "นักเรียนทั้งหมด", "count": student_count, "unit": "คน", "color": PRIMARY},
            {"label": "ครูทั้งหมด", "count": teacher_count, "unit": "คน", "color": SUCCESS},
            {"label": "ห้องเรียน", "count": class_count, "unit": "ห้อง", "color": WARNING},
        ]

        for idx, item in enumerate(summary_items):
            card = ctk.CTkFrame(
                cards_frame,
                fg_color=SURFACE,
                corner_radius=RADIUS_CARD,
                border_width=1,
                border_color=TABLE_BORDER,
            )
            card.grid(row=0, column=idx, padx=(0 if idx == 0 else S, 0 if idx == 2 else S), pady=0, sticky="nsew")

            # Card interior layout
            card_inner = ctk.CTkFrame(card, fg_color="transparent")
            card_inner.pack(fill="both", expand=True, padx=M, pady=M)

            # Left: accent color strip
            accent = ctk.CTkFrame(
                card_inner,
                fg_color=item["color"],
                width=4,
                corner_radius=2,
            )
            accent.pack(side="left", fill="y", padx=(0, S))

            # Text block
            text_block = ctk.CTkFrame(card_inner, fg_color="transparent")
            text_block.pack(side="left", fill="both", expand=True)

            ctk.CTkLabel(
                text_block,
                text=item["label"],
                font=ctk.CTkFont(family="TH Sarabun New", size=14),
                text_color=TEXT_CAPTION,
                anchor="w",
            ).pack(anchor="w")

            # Count number - H1 28px
            count_frame = ctk.CTkFrame(text_block, fg_color="transparent")
            count_frame.pack(anchor="w")

            ctk.CTkLabel(
                count_frame,
                text=str(item["count"]),
                font=ctk.CTkFont(family="TH Sarabun New", size=22, weight="bold"),
                text_color=item["color"],
            ).pack(side="left")

            ctk.CTkLabel(
                count_frame,
                text=f" {item['unit']}",
                font=ctk.CTkFont(family="TH Sarabun New", size=14),
                text_color=TEXT_CAPTION,
            ).pack(side="left", pady=(S, 0))

    def _create_report_card(self, parent, report, row, col):
        """สร้าง Report Card แต่ละใบ"""

        accent_color = CARD_COLORS.get(report["key"], NEUTRAL)

        # Card frame - radius 12px
        card = ctk.CTkFrame(
            parent,
            fg_color=SURFACE,
            corner_radius=RADIUS_CARD,
            border_width=1,
            border_color=TABLE_BORDER,
        )
        card.grid(
            row=row, column=col,
            padx=(0 if col == 0 else S, 0 if col == 2 else S),
            pady=S,
            sticky="nsew",
        )

        # Card content
        content = ctk.CTkFrame(card, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=L, pady=L)

        # Top row: icon + title
        top_row = ctk.CTkFrame(content, fg_color="transparent")
        top_row.pack(fill="x", pady=(0, M))

        # Icon circle - 48px
        icon_frame = ctk.CTkFrame(
            top_row,
            fg_color=accent_color,
            width=48,
            height=48,
            corner_radius=RADIUS_BUTTON,
        )
        icon_frame.pack(side="left")
        icon_frame.pack_propagate(False)

        icon_image = IconManager.get(report["icon"], 24, color="#FFFFFF", dark_color="#FFFFFF")
        if icon_image:
            ctk.CTkLabel(
                icon_frame,
                text="",
                image=icon_image,
            ).pack(expand=True)
        else:
            ctk.CTkLabel(
                icon_frame,
                text=report["icon"][0].upper(),
                font=ctk.CTkFont(family="TH Sarabun New", size=20, weight="bold"),
                text_color="#FFFFFF",
            ).pack(expand=True)

        # Title + desc
        title_block = ctk.CTkFrame(top_row, fg_color="transparent")
        title_block.pack(side="left", padx=(M, 0), fill="x", expand=True)

        ctk.CTkLabel(
            title_block,
            text=report["title"],
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            text_color=TEXT_H2,
            anchor="w",
        ).pack(anchor="w")

        ctk.CTkLabel(
            title_block,
            text=report["desc"],
            font=ctk.CTkFont(family="TH Sarabun New", size=12),
            text_color=TEXT_CAPTION,
            anchor="w",
        ).pack(anchor="w")

        # Separator line
        sep = ctk.CTkFrame(content, fg_color=TABLE_BORDER, height=1)
        sep.pack(fill="x", pady=(0, M))

        # Export buttons - SECONDARY style (outlined)
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x")

        # Excel button - SECONDARY outlined
        ctk.CTkButton(
            btn_frame,
            text="Excel",
            command=report["excel"],
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            width=0,
            height=36,
            corner_radius=RADIUS_BUTTON,
            fg_color="transparent",
            border_width=1,
            border_color=accent_color,
            text_color=accent_color,
            hover_color=TABLE_HOVER,
            image=IconManager.get("file-export", 14, color=accent_color, dark_color=accent_color),
            compound="left",
        ).pack(side="left", fill="x", expand=True, padx=(0, XS))

        # PDF button - SECONDARY outlined
        ctk.CTkButton(
            btn_frame,
            text="PDF",
            command=report["pdf"],
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            width=0,
            height=36,
            corner_radius=RADIUS_BUTTON,
            fg_color="transparent",
            border_width=1,
            border_color=accent_color,
            text_color=accent_color,
            hover_color=TABLE_HOVER,
            image=IconManager.get("file-pdf", 14, color=accent_color, dark_color=accent_color),
            compound="left",
        ).pack(side="left", fill="x", expand=True, padx=(XS, 0))

    # ==================== EXPORT FUNCTIONS ====================

    def export_students_excel(self):
        """Export รายชื่อนักเรียนเป็น Excel"""

        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"รายชื่อนักเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "รายชื่อนักเรียน"

            # Style - design system colors (blue header, white text)
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color="E5E7EB"),
                right=Side(style='thin', color="E5E7EB"),
                top=Side(style='thin', color="E5E7EB"),
                bottom=Side(style='thin', color="E5E7EB"),
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            # Header row
            headers = ["รหัสนักเรียน", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา", "วันเกิด", "ผู้ปกครอง", "เบอร์ติดต่อ"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Data rows
            for row_idx, student in enumerate(students, start=2):
                data = [
                    student['student_id'],
                    student['title'],
                    student['first_name'],
                    student['last_name'],
                    student['class_room'],
                    student['class_year'],
                    student['birth_date'] or "-",
                    student['parent_name'] or "-",
                    student['parent_phone'] or "-",
                ]

                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx in [1, 2, 5, 6] else 'left')
                    # Stripe pattern
                    if row_idx % 2 == 0:
                        cell.fill = stripe_fill

            # Column widths
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 15

            wb.save(file_path)
            self.update_status(f"Export รายชื่อนักเรียน {len(students)} รายการเป็น Excel สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_students_pdf(self):
        """Export รายชื่อนักเรียนเป็น PDF"""

        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"รายชื่อนักเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []

            # Title
            styles = getSampleStyleSheet()
            title = Paragraph("รายชื่อนักเรียนทั้งหมด", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=1,
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            # Table data
            data = [["รหัส", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา"]]

            for s in students:
                data.append([
                    s['student_id'],
                    s['title'],
                    s['first_name'],
                    s['last_name'],
                    s['class_room'],
                    s['class_year'],
                ])

            table = Table(data, colWidths=[3 * cm, 2 * cm, 4 * cm, 4 * cm, 2.5 * cm, 2.5 * cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#111827")),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))

            elements.append(table)
            doc.build(elements)

            self.update_status(f"Export รายชื่อนักเรียน {len(students)} รายการเป็น PDF สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_attendance_excel(self):
        """Export การเช็คชื่อเป็น Excel"""
        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"สถิติการเช็คชื่อ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "สถิติการเช็คชื่อ"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            headers = ["รหัส", "ชื่อ-สกุล", "ห้อง", "มา", "ขาด", "ลา", "มาสาย", "รวม"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for row_idx, student in enumerate(students, start=2):
                stats = self.db.get_attendance_stats(student['student_id'])
                total = stats['มา'] + stats['ขาด'] + stats['ลา'] + stats['มาสาย']
                full_name = f"{student['title']}{student['first_name']} {student['last_name']}"
                data = [
                    student['student_id'],
                    full_name,
                    student['class_room'],
                    stats['มา'],
                    stats['ขาด'],
                    stats['ลา'],
                    stats['มาสาย'],
                    total,
                ]
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left')
                    if row_idx % 2 == 0:
                        cell.fill = stripe_fill

            col_widths = {'A': 12, 'B': 24, 'C': 10, 'D': 8, 'E': 8, 'F': 8, 'G': 10, 'H': 8}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            self.update_status(f"Export สถิติการเช็คชื่อ {len(students)} รายการเป็น Excel สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_attendance_pdf(self):
        """Export การเช็คชื่อเป็น PDF"""
        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"สถิติการเช็คชื่อ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            title = Paragraph("สถิติการเช็คชื่อนักเรียน", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=1,
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            data = [["รหัส", "ชื่อ-สกุล", "ห้อง", "มา", "ขาด", "ลา", "มาสาย", "รวม"]]
            for student in students:
                stats = self.db.get_attendance_stats(student['student_id'])
                total = stats['มา'] + stats['ขาด'] + stats['ลา'] + stats['มาสาย']
                full_name = f"{student['title']}{student['first_name']} {student['last_name']}"
                data.append([
                    student['student_id'],
                    full_name,
                    student['class_room'],
                    str(stats['มา']),
                    str(stats['ขาด']),
                    str(stats['ลา']),
                    str(stats['มาสาย']),
                    str(total),
                ])

            table = Table(data, colWidths=[2.5*cm, 5*cm, 2*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.8*cm, 1.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))

            elements.append(table)
            doc.build(elements)
            self.update_status(f"Export สถิติการเช็คชื่อ {len(students)} รายการเป็น PDF สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_health_excel(self):
        """Export ข้อมูลสุขภาพเป็น Excel"""
        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"ข้อมูลสุขภาพ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ข้อมูลสุขภาพ"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            headers = ["รหัส", "ชื่อ-สกุล", "ห้อง", "น้ำหนัก(kg)", "ส่วนสูง(cm)", "BMI", "สถานะ"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for row_idx, student in enumerate(students, start=2):
                health = self.db.get_latest_health(student['student_id'])
                full_name = f"{student['title']}{student['first_name']} {student['last_name']}"
                if health:
                    weight = health['weight_kg'] if health['weight_kg'] is not None else "-"
                    height = health['height_cm'] if health['height_cm'] is not None else "-"
                    bmi_val = health['bmi'] if health['bmi'] is not None else "-"
                    if isinstance(bmi_val, (int, float)):
                        bmi_str = f"{bmi_val:.2f}"
                        if bmi_val < 18.5:
                            status = "น้ำหนักต่ำกว่าเกณฑ์"
                        elif bmi_val < 23:
                            status = "น้ำหนักปกติ"
                        elif bmi_val < 25:
                            status = "น้ำหนักเกิน"
                        else:
                            status = "อ้วน"
                    else:
                        bmi_str = "-"
                        status = "-"
                else:
                    weight = "-"
                    height = "-"
                    bmi_str = "-"
                    status = "-"

                data = [
                    student['student_id'],
                    full_name,
                    student['class_room'],
                    weight,
                    height,
                    bmi_str,
                    status,
                ]
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left')
                    if row_idx % 2 == 0:
                        cell.fill = stripe_fill

            col_widths = {'A': 12, 'B': 24, 'C': 10, 'D': 14, 'E': 14, 'F': 10, 'G': 20}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            self.update_status(f"Export ข้อมูลสุขภาพ {len(students)} รายการเป็น Excel สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_health_pdf(self):
        """Export ข้อมูลสุขภาพเป็น PDF"""
        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"ข้อมูลสุขภาพ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            title = Paragraph("ข้อมูลสุขภาพนักเรียน", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=1,
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            data = [["รหัส", "ชื่อ-สกุล", "ห้อง", "น้ำหนัก(kg)", "ส่วนสูง(cm)", "BMI", "สถานะ"]]
            for student in students:
                health = self.db.get_latest_health(student['student_id'])
                full_name = f"{student['title']}{student['first_name']} {student['last_name']}"
                if health:
                    weight = f"{health['weight_kg']:.1f}" if health['weight_kg'] is not None else "-"
                    height = f"{health['height_cm']:.1f}" if health['height_cm'] is not None else "-"
                    bmi_val = health['bmi']
                    if bmi_val is not None:
                        bmi_str = f"{bmi_val:.2f}"
                        if bmi_val < 18.5:
                            status = "ต่ำกว่าเกณฑ์"
                        elif bmi_val < 23:
                            status = "ปกติ"
                        elif bmi_val < 25:
                            status = "น้ำหนักเกิน"
                        else:
                            status = "อ้วน"
                    else:
                        bmi_str = "-"
                        status = "-"
                else:
                    weight = "-"
                    height = "-"
                    bmi_str = "-"
                    status = "-"

                data.append([
                    student['student_id'],
                    full_name,
                    student['class_room'],
                    weight,
                    height,
                    bmi_str,
                    status,
                ])

            table = Table(data, colWidths=[2.5*cm, 4.5*cm, 1.8*cm, 2.2*cm, 2.2*cm, 1.8*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))

            elements.append(table)
            doc.build(elements)
            self.update_status(f"Export ข้อมูลสุขภาพ {len(students)} รายการเป็น PDF สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_grades_excel(self):
        """Export เกรดเป็น Excel"""
        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"ผลการเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ผลการเรียน"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            headers = ["รหัส", "ชื่อ-สกุล", "ห้อง", "รหัสวิชา", "ชื่อวิชา", "คะแนน", "เกรด", "ปีการศึกษา", "ภาคเรียน"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            row_idx = 2
            total_rows = 0
            for student in students:
                grades = self.db.get_grades(student['student_id'])
                if not grades:
                    continue
                full_name = f"{student['title']}{student['first_name']} {student['last_name']}"
                for grade in grades:
                    data = [
                        student['student_id'],
                        full_name,
                        student['class_room'],
                        grade['subject_code'],
                        grade['subject_name'],
                        grade['score'] if grade['score'] is not None else "-",
                        grade['grade'] if grade['grade'] else "-",
                        grade['academic_year'],
                        grade['semester'],
                    ]
                    for col_idx, value in enumerate(data, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center' if col_idx not in [2, 5] else 'left')
                        if row_idx % 2 == 0:
                            cell.fill = stripe_fill
                    row_idx += 1
                    total_rows += 1

            if total_rows == 0:
                messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลเกรด")
                return

            col_widths = {'A': 12, 'B': 24, 'C': 10, 'D': 12, 'E': 24, 'F': 10, 'G': 8, 'H': 14, 'I': 10}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            self.update_status(f"Export ผลการเรียน {total_rows} รายการเป็น Excel สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_grades_pdf(self):
        """Export เกรดเป็น PDF"""
        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"ผลการเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            title = Paragraph("ผลการเรียนนักเรียน", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=1,
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            data = [["รหัส", "ชื่อ-สกุล", "ห้อง", "รหัสวิชา", "ชื่อวิชา", "คะแนน", "เกรด", "ปีการศึกษา", "ภาคเรียน"]]
            total_rows = 0
            for student in students:
                grades = self.db.get_grades(student['student_id'])
                if not grades:
                    continue
                full_name = f"{student['title']}{student['first_name']} {student['last_name']}"
                for grade in grades:
                    data.append([
                        student['student_id'],
                        full_name,
                        student['class_room'],
                        grade['subject_code'],
                        grade['subject_name'],
                        str(grade['score']) if grade['score'] is not None else "-",
                        grade['grade'] if grade['grade'] else "-",
                        grade['academic_year'],
                        grade['semester'],
                    ])
                    total_rows += 1

            if total_rows == 0:
                messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลเกรด")
                return

            table = Table(data, colWidths=[2*cm, 4*cm, 1.5*cm, 2*cm, 3.5*cm, 1.8*cm, 1.5*cm, 2*cm, 1.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))

            elements.append(table)
            doc.build(elements)
            self.update_status(f"Export ผลการเรียน {total_rows} รายการเป็น PDF สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_schedule_excel(self):
        """Export ตารางเรียนเป็น Excel"""
        schedules = self.db.get_all_schedules()
        if not schedules:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลตารางเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"ตารางเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ตารางเรียน"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            headers = ["ห้อง", "วัน", "คาบ", "วิชา", "ครู", "เวลาเริ่ม", "เวลาสิ้นสุด"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for row_idx, sched in enumerate(schedules, start=2):
                teacher_name = f"{sched.get('title', '')}{sched.get('first_name', '')} {sched.get('last_name', '')}".strip()
                data = [
                    sched['class_room'],
                    sched['day_of_week'],
                    sched['period_no'],
                    sched['subject_name'],
                    teacher_name,
                    sched.get('start_time') or "-",
                    sched.get('end_time') or "-",
                ]
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx not in [4, 5] else 'left')
                    if row_idx % 2 == 0:
                        cell.fill = stripe_fill

            col_widths = {'A': 10, 'B': 14, 'C': 6, 'D': 24, 'E': 22, 'F': 12, 'G': 14}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            self.update_status(f"Export ตารางเรียน {len(schedules)} รายการเป็น Excel สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_schedule_pdf(self):
        """Export ตารางเรียนเป็น PDF"""
        schedules = self.db.get_all_schedules()
        if not schedules:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลตารางเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"ตารางเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            title = Paragraph("ตารางเรียน/ตารางสอน", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=1,
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            data = [["ห้อง", "วัน", "คาบ", "วิชา", "ครู", "เวลาเริ่ม", "เวลาสิ้นสุด"]]
            for sched in schedules:
                teacher_name = f"{sched.get('title', '')}{sched.get('first_name', '')} {sched.get('last_name', '')}".strip()
                data.append([
                    sched['class_room'],
                    sched['day_of_week'],
                    str(sched['period_no']),
                    sched['subject_name'],
                    teacher_name,
                    sched.get('start_time') or "-",
                    sched.get('end_time') or "-",
                ])

            table = Table(data, colWidths=[2*cm, 2.5*cm, 1.2*cm, 4.5*cm, 4*cm, 2*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (3, 1), (4, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))

            elements.append(table)
            doc.build(elements)
            self.update_status(f"Export ตารางเรียน {len(schedules)} รายการเป็น PDF สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_all_excel(self):
        """Export ข้อมูลทั้งหมดเป็น Excel (Multiple Sheets)"""

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"สรุปข้อมูลทั้งหมด_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()

            # Style
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color="E5E7EB"),
                right=Side(style='thin', color="E5E7EB"),
                top=Side(style='thin', color="E5E7EB"),
                bottom=Side(style='thin', color="E5E7EB"),
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            # ========== Sheet 1: Students ==========
            ws1 = wb.active
            ws1.title = "รายชื่อนักเรียน"

            students = self.db.get_all_students()
            headers1 = ["รหัส", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา"]
            for col, h in enumerate(headers1, start=1):
                cell = ws1.cell(row=1, column=col)
                cell.value = h
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            for row, s in enumerate(students, start=2):
                values = [s['student_id'], s['title'], s['first_name'], s['last_name'], s['class_room'], s['class_year']]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws1.cell(row=row, column=col_idx)
                    cell.value = val
                    cell.border = border
                    if row % 2 == 0:
                        cell.fill = stripe_fill

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws1.column_dimensions[col_letter].width = 15

            # ========== Sheet 2: Teachers ==========
            ws2 = wb.create_sheet("ครู")
            teachers = self.db.get_all_teachers()
            headers2 = ["รหัสครู", "คำนำหน้า", "ชื่อ", "นามสกุล", "เบอร์ติดต่อ"]
            for col, h in enumerate(headers2, start=1):
                cell = ws2.cell(row=1, column=col)
                cell.value = h
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            for row, t in enumerate(teachers, start=2):
                values = [t['teacher_id'], t['title'], t['first_name'], t['last_name'], t['phone'] or "-"]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws2.cell(row=row, column=col_idx)
                    cell.value = val
                    cell.border = border
                    if row % 2 == 0:
                        cell.fill = stripe_fill

            for col_letter in ['A', 'B', 'C', 'D', 'E']:
                ws2.column_dimensions[col_letter].width = 15

            # ========== Sheet 3: Stats ==========
            ws3 = wb.create_sheet("สถิติ")

            stats_headers = ["รายการ", "จำนวน"]
            for col, h in enumerate(stats_headers, start=1):
                cell = ws3.cell(row=1, column=col)
                cell.value = h
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            class_rooms = self.db.get_class_rooms()
            stats_data = [
                ("จำนวนนักเรียน", f"{len(students)} คน"),
                ("จำนวนครู", f"{len(teachers)} คน"),
                ("จำนวนห้องเรียน", f"{len(class_rooms)} ห้อง"),
            ]
            for row_idx, (label, value) in enumerate(stats_data, start=2):
                cell_label = ws3.cell(row=row_idx, column=1)
                cell_label.value = label
                cell_label.border = border

                cell_value = ws3.cell(row=row_idx, column=2)
                cell_value.value = value
                cell_value.border = border
                cell_value.alignment = Alignment(horizontal='center')

                if row_idx % 2 == 0:
                    cell_label.fill = stripe_fill
                    cell_value.fill = stripe_fill

            ws3.column_dimensions['A'].width = 20
            ws3.column_dimensions['B'].width = 15

            wb.save(file_path)
            self.update_status("Export สรุปข้อมูลทั้งหมดเป็น Excel สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_all_pdf(self):
        """Export สรุปข้อมูลทั้งหมดเป็น PDF"""

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"สรุปข้อมูลทั้งหมด_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title = Paragraph("สรุปข้อมูลระบบบริหารจัดการโรงเรียน", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=20,
                alignment=1,
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            # Stats data
            students = self.db.get_all_students()
            teachers = self.db.get_all_teachers()
            class_rooms = self.db.get_class_rooms()

            stats_data = [
                ["รายการ", "จำนวน"],
                ["นักเรียนทั้งหมด", f"{len(students)} คน"],
                ["ครูทั้งหมด", f"{len(teachers)} คน"],
                ["ห้องเรียน", f"{len(class_rooms)} ห้อง"],
            ]

            stats_table = Table(stats_data, colWidths=[10 * cm, 5 * cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#111827")),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))

            elements.append(stats_table)
            elements.append(Spacer(1, 1 * cm))

            # Footer info
            info = Paragraph(
                f"รายงานนี้สร้างเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                ParagraphStyle('Info', parent=styles['Normal'], fontName=font_name, fontSize=10, alignment=1)
            )
            elements.append(info)

            doc.build(elements)

            self.update_status("Export สรุปข้อมูลเป็น PDF สำเร็จ", "success")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")
