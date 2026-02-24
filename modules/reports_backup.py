"""
modules/reports.py
โมดูลรายงาน
- Export ทุกโมดูลเป็น Excel/PDF
- พิมพ์โดยตรง
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class ReportsModule:
    """โมดูลรายงาน"""

    def __init__(self, parent, db, update_status_callback):
        self.parent = parent
        self.db = db
        self.update_status = update_status_callback

        # สร้าง UI
        self.create_ui()

    def create_ui(self):
        """สร้าง UI ของโมดูล"""

        # Main frame
        main_frame = ctk.CTkFrame(self.parent, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Title
        ctk.CTkLabel(
            main_frame,
            text="📊 รายงานและส่งออกข้อมูล",
            font=ctk.CTkFont(family="TH Sarabun New", size=24, weight="bold")
        ).pack(pady=30)

        # Grid สำหรับปุ่ม
        button_grid = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_grid.pack(expand=True)

        # รายงานแต่ละประเภท
        reports = [
            {
                "title": "รายชื่อนักเรียน",
                "icon": "👨‍🎓",
                "color": "#3498DB",
                "excel": self.export_students_excel,
                "pdf": self.export_students_pdf
            },
            {
                "title": "การเช็คชื่อ",
                "icon": "✓",
                "color": "#27AE60",
                "excel": self.export_attendance_excel,
                "pdf": self.export_attendance_pdf
            },
            {
                "title": "ข้อมูลสุขภาพ",
                "icon": "💊",
                "color": "#E74C3C",
                "excel": self.export_health_excel,
                "pdf": self.export_health_pdf
            },
            {
                "title": "เกรดและผลการเรียน",
                "icon": "📝",
                "color": "#F39C12",
                "excel": self.export_grades_excel,
                "pdf": self.export_grades_pdf
            },
            {
                "title": "ตารางเรียน/สอน",
                "icon": "📅",
                "color": "#8E44AD",
                "excel": self.export_schedule_excel,
                "pdf": self.export_schedule_pdf
            },
            {
                "title": "สรุปข้อมูลทั้งหมด",
                "icon": "📊",
                "color": "#1F4E78",
                "excel": self.export_all_excel,
                "pdf": self.export_all_pdf
            }
        ]

        # สร้างปุ่มในรูป Grid 2 คอลัมน์
        for idx, report in enumerate(reports):
            row = idx // 2
            col = idx % 2

            # กรอบแต่ละรายงาน
            report_frame = ctk.CTkFrame(
                button_grid,
                fg_color=report["color"],
                corner_radius=15
            )
            report_frame.grid(row=row, column=col, padx=20, pady=20, sticky="nsew")

            # ไอคอนและชื่อ
            ctk.CTkLabel(
                report_frame,
                text=report["icon"],
                font=ctk.CTkFont(size=48)
            ).pack(pady=(20, 10))

            ctk.CTkLabel(
                report_frame,
                text=report["title"],
                font=ctk.CTkFont(family="TH Sarabun New", size=18, weight="bold"),
                text_color="white"
            ).pack(pady=(0, 20))

            # ปุ่ม Export
            btn_frame = ctk.CTkFrame(report_frame, fg_color="transparent")
            btn_frame.pack(pady=(0, 20))

            ctk.CTkButton(
                btn_frame,
                text="📊 Excel",
                command=report["excel"],
                font=ctk.CTkFont(family="TH Sarabun New", size=14),
                width=120,
                height=35,
                fg_color="white",
                text_color=report["color"],
                hover_color="#ECF0F1"
            ).pack(side="left", padx=5)

            ctk.CTkButton(
                btn_frame,
                text="📄 PDF",
                command=report["pdf"],
                font=ctk.CTkFont(family="TH Sarabun New", size=14),
                width=120,
                height=35,
                fg_color="white",
                text_color=report["color"],
                hover_color="#ECF0F1"
            ).pack(side="left", padx=5)

        # ปรับ column weight
        button_grid.columnconfigure(0, weight=1)
        button_grid.columnconfigure(1, weight=1)

    # ==================== EXPORT FUNCTIONS ====================

    def export_students_excel(self):
        """Export รายชื่อนักเรียนเป็น Excel"""

        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"รายชื่อนักเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "รายชื่อนักเรียน"

            # สไตล์
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # หัวตาราง
            headers = ["รหัสนักเรียน", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา", "วันเกิด", "ผู้ปกครอง", "เบอร์ติดต่อ"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # ข้อมูล
            for row_idx, student in enumerate(students, start=2):
                data = [
                    student['student_id'],
                    student['title'],
                    student['first_name'],
                    student['last_name'],
                    student['class_room'],
                    student['class_year'],
                    student['birth_date'] or "-",
                    student['parent_name'] or "-",
                    student['parent_phone'] or "-"
                ]

                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx in [1, 2, 5, 6] else 'left')

            # ปรับความกว้างคอลัมน์
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 15

            wb.save(file_path)
            self.update_status("Export รายชื่อนักเรียนเป็น Excel สำเร็จ")
            messagebox.showinfo("สำเร็จ", f"Export {len(students)} รายการเรียบร้อย")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_students_pdf(self):
        """Export รายชื่อนักเรียนเป็น PDF"""

        students = self.db.get_all_students()
        if not students:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลนักเรียน")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"รายชื่อนักเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not file_path:
            return

        try:
            # ฟอนต์
            try:
                pdfmetrics.registerFont(TTFont('Sarabun', 'THSarabunNew.ttf'))
                font_name = 'Sarabun'
            except:
                font_name = 'Helvetica'

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []

            # Title
            styles = getSampleStyleSheet()
            title = Paragraph("รายชื่อนักเรียนทั้งหมด", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=1
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5*cm))

            # Table
            data = [["รหัส", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา"]]

            for s in students:
                data.append([
                    s['student_id'],
                    s['title'],
                    s['first_name'],
                    s['last_name'],
                    s['class_room'],
                    s['class_year']
                ])

            table = Table(data, colWidths=[3*cm, 2*cm, 4*cm, 4*cm, 2.5*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))

            elements.append(table)
            doc.build(elements)

            self.update_status("Export PDF สำเร็จ")
            messagebox.showinfo("สำเร็จ", f"Export {len(students)} รายการเรียบร้อย")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_attendance_excel(self):
        """Export การเช็คชื่อเป็น Excel"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export การเช็คชื่อกำลังพัฒนา\nสามารถ Export ได้จากแท็บเช็คชื่อ")

    def export_attendance_pdf(self):
        """Export การเช็คชื่อเป็น PDF"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export การเช็คชื่อกำลังพัฒนา\nสามารถ Export ได้จากแท็บเช็คชื่อ")

    def export_health_excel(self):
        """Export ข้อมูลสุขภาพเป็น Excel"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export ข้อมูลสุขภาพกำลังพัฒนา\nสามารถ Export ได้จากแท็บสุขภาพ")

    def export_health_pdf(self):
        """Export ข้อมูลสุขภาพเป็น PDF"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export ข้อมูลสุขภาพกำลังพัฒนา\nสามารถ Export ได้จากแท็บสุขภาพ")

    def export_grades_excel(self):
        """Export เกรดเป็น Excel"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export เกรดกำลังพัฒนา\nสามารถ Export Transcript ได้จากแท็บบันทึกเกรด")

    def export_grades_pdf(self):
        """Export เกรดเป็น PDF"""
        messagebox.showinfo("ข้อมูล", "สามารถ Export Transcript ได้จากแท็บบันทึกเกรด")

    def export_schedule_excel(self):
        """Export ตารางเรียนเป็น Excel"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export ตารางเรียนกำลังพัฒนา")

    def export_schedule_pdf(self):
        """Export ตารางเรียนเป็น PDF"""
        messagebox.showinfo("ข้อมูล", "ฟีเจอร์ Export ตารางเรียนกำลังพัฒนา")

    def export_all_excel(self):
        """Export ข้อมูลทั้งหมดเป็น Excel (Multiple Sheets)"""

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"สรุปข้อมูลทั้งหมด_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()

            # Sheet 1: รายชื่อนักเรียน
            ws1 = wb.active
            ws1.title = "รายชื่อนักเรียน"

            students = self.db.get_all_students()
            headers1 = ["รหัส", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา"]
            for col, h in enumerate(headers1, start=1):
                ws1.cell(row=1, column=col).value = h

            for row, s in enumerate(students, start=2):
                ws1.cell(row=row, column=1).value = s['student_id']
                ws1.cell(row=row, column=2).value = s['title']
                ws1.cell(row=row, column=3).value = s['first_name']
                ws1.cell(row=row, column=4).value = s['last_name']
                ws1.cell(row=row, column=5).value = s['class_room']
                ws1.cell(row=row, column=6).value = s['class_year']

            # Sheet 2: ครู
            ws2 = wb.create_sheet("ครู")
            teachers = self.db.get_all_teachers()
            headers2 = ["รหัสครู", "คำนำหน้า", "ชื่อ", "นามสกุล", "เบอร์ติดต่อ"]
            for col, h in enumerate(headers2, start=1):
                ws2.cell(row=1, column=col).value = h

            for row, t in enumerate(teachers, start=2):
                ws2.cell(row=row, column=1).value = t['teacher_id']
                ws2.cell(row=row, column=2).value = t['title']
                ws2.cell(row=row, column=3).value = t['first_name']
                ws2.cell(row=row, column=4).value = t['last_name']
                ws2.cell(row=row, column=5).value = t['phone'] or "-"

            # Sheet 3: สถิติ
            ws3 = wb.create_sheet("สถิติ")
            ws3.cell(row=1, column=1).value = "สถิติข้อมูล"
            ws3.cell(row=2, column=1).value = "จำนวนนักเรียน"
            ws3.cell(row=2, column=2).value = len(students)
            ws3.cell(row=3, column=1).value = "จำนวนครู"
            ws3.cell(row=3, column=2).value = len(teachers)
            ws3.cell(row=4, column=1).value = "จำนวนห้องเรียน"
            ws3.cell(row=4, column=2).value = len(self.db.get_class_rooms())

            wb.save(file_path)
            self.update_status("Export สรุปข้อมูลทั้งหมดเป็น Excel สำเร็จ")
            messagebox.showinfo("สำเร็จ", "Export ข้อมูลทั้งหมดเรียบร้อย")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_all_pdf(self):
        """Export สรุปข้อมูลทั้งหมดเป็น PDF"""

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"สรุปข้อมูลทั้งหมด_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not file_path:
            return

        try:
            # ฟอนต์
            try:
                pdfmetrics.registerFont(TTFont('Sarabun', 'THSarabunNew.ttf'))
                font_name = 'Sarabun'
            except:
                font_name = 'Helvetica'

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title = Paragraph("สรุปข้อมูลระบบบริหารจัดการโรงเรียน", ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=20,
                alignment=1
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5*cm))

            # สถิติ
            students = self.db.get_all_students()
            teachers = self.db.get_all_teachers()
            class_rooms = self.db.get_class_rooms()

            stats_data = [
                ["รายการ", "จำนวน"],
                ["นักเรียนทั้งหมด", f"{len(students)} คน"],
                ["ครูทั้งหมด", f"{len(teachers)} คน"],
                ["ห้องเรียน", f"{len(class_rooms)} ห้อง"]
            ]

            stats_table = Table(stats_data, colWidths=[10*cm, 5*cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
            ]))

            elements.append(stats_table)
            elements.append(Spacer(1, 1*cm))

            # ข้อมูลเพิ่มเติม
            info = Paragraph(
                f"รายงานนี้สร้างเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                ParagraphStyle('Info', parent=styles['Normal'], fontName=font_name, fontSize=10, alignment=1)
            )
            elements.append(info)

            doc.build(elements)

            self.update_status("Export PDF สำเร็จ")
            messagebox.showinfo("สำเร็จ", "Export สรุปข้อมูลเรียบร้อย")

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")
