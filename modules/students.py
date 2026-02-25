"""
modules/students.py
โมดูลจัดการนักเรียน - Design System v3.0
- ตาราง striped + hover ตาม design system
- ปุ่ม PRIMARY ไม่เกิน 1 ปุ่มต่อ section
- Form modal กว้างไม่เกิน 560px, radius 16px
- Validation แสดง error ใต้ field สีแดง
- Empty state แสดงเมื่อไม่มีข้อมูล
- Toast notification สีเขียว 3 วินาที
"""

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from modules.icons import IconManager
from modules.pdf_utils import get_thai_font

# ==================== Design System v4.0 ====================
# Accent Colors (10%) - Updated
PRIMARY = "#3B82F6"   # Blue
SUCCESS = "#10B981"   # Green
WARNING = "#F59E0B"   # Amber
DANGER = "#EF4444"    # Red
NEUTRAL = "#64748B"   # Slate

# Typography Colors - Better contrast
TEXT_H1 = "#0F172A"    # Near black
TEXT_H2 = "#1E293B"   # Dark slate
TEXT_H3 = "#334155"   # Medium slate
TEXT_BODY = "#475569"  # Body
TEXT_CAPTION = "#94A3B8"  # Muted

# Table - Modern
TABLE_HEADER_BG = "#F1F5F9"  # Slate-100
TABLE_HOVER = "#E0F2FE"      # Light blue
TABLE_STRIPE = "#F8FAFC"    # Very light
TABLE_BORDER = "#E2E8F0"   # Light border

# Spacing
XS, S, M, L, XL, XXL = 4, 8, 16, 24, 32, 48

# Radius - Softer
RADIUS_BUTTON = 8
RADIUS_CARD = 12
RADIUS_MODAL = 16
RADIUS_INPUT = 8

# Input - Modern
INPUT_BORDER = "#CBD5E1"  # Slate-300
INPUT_FOCUS = PRIMARY


class StudentsModule:
    """โมดูลจัดการนักเรียน - Design System v3.0"""

    def __init__(self, parent, db, update_status_callback):
        self.parent = parent
        self.db = db
        self.update_status = update_status_callback
        self.students_data = []
        self.selected_student_id = None

        self.create_ui()
        self.setup_table_style()
        self.load_students()

    def create_ui(self):
        """สร้าง UI ตาม Design System
        ใช้ CTkScrollableFrame ครอบเนื้อหาทั้งหมด เพื่อให้เลื่อนดูได้เมื่อหน้าจอเล็ก
        """

        # CTkScrollableFrame ครอบเนื้อหาทั้งหมด - เลื่อนดูได้เมื่อเนื้อหาล้น
        self.content_frame = ctk.CTkScrollableFrame(
            self.parent,
            fg_color="transparent",
            scrollbar_button_color="#CBD5E1",
            scrollbar_button_hover_color=PRIMARY
        )
        self.content_frame.pack(fill="both", expand=True)

        # === ส่วนค้นหาและกรอง (card style) ===
        search_card = ctk.CTkFrame(
            self.content_frame, fg_color="#FFFFFF",
            corner_radius=RADIUS_CARD, border_width=1, border_color="#E2E8F0"
        )
        search_card.pack(fill="x", padx=L, pady=(L, M))

        search_frame = ctk.CTkFrame(search_card, fg_color="transparent")
        search_frame.pack(fill="x", padx=M, pady=M)

        # กล่องค้นหา (with icon)
        self.search_var = ctk.StringVar()
        self.search_var.trace("w", lambda *args: self.search_students())
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.search_var,
            placeholder_text="🔍 ค้นหาชื่อ, นามสกุล หรือรหัส...",
            width=320,
            height=42,
            font=ctk.CTkFont(family="Kanit", size=14),
            corner_radius=RADIUS_INPUT,
            border_width=1,
            border_color=INPUT_BORDER,
            placeholder_text_color="#94A3B8"
        )
        search_entry.pack(side="left", padx=(0, M))

        # กรองห้อง
        ctk.CTkLabel(
            search_frame,
            text="ห้อง:",
            font=("Kanit", 14),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, S))

        self.class_var = ctk.StringVar(value="ทั้งหมด")
        class_options = ["ทั้งหมด"] + self.db.get_class_rooms()
        ctk.CTkOptionMenu(
            search_frame,
            variable=self.class_var,
            values=class_options,
            command=lambda x: self.load_students(),
            width=140,
            height=38,
            font=ctk.CTkFont(family="Kanit", size=13),
            corner_radius=RADIUS_INPUT,
            fg_color="#F8FAFC",
            button_color=PRIMARY,
            button_hover_color="#2563EB",
            text_color=TEXT_H1,
            dropdown_fg_color="#FFFFFF",
            dropdown_hover_color="#E0F2FE",
            dropdown_text_color=TEXT_H1,
            dropdown_font=ctk.CTkFont(family="Kanit", size=13)
        ).pack(side="left", padx=(0, M))

        # กรองปีการศึกษา
        ctk.CTkLabel(
            search_frame,
            text="ปี:",
            font=("Kanit", 14),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, S))

        self.year_var = ctk.StringVar(value="ทั้งหมด")
        year_options = ["ทั้งหมด"] + self.db.get_class_years()
        ctk.CTkOptionMenu(
            search_frame,
            variable=self.year_var,
            values=year_options,
            command=lambda x: self.load_students(),
            width=110,
            height=38,
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            corner_radius=20,
            fg_color="#EFF6FF",
            button_color="#EFF6FF",
            button_hover_color="#DBEAFE",
            text_color="#1E40AF",
            dropdown_fg_color="#FFFFFF",
            dropdown_hover_color="#E0F2FE",
            dropdown_text_color=TEXT_H1,
            dropdown_font=ctk.CTkFont(family="Kanit", size=13)
        ).pack(side="left")

        # ปุ่ม Refresh - Modern style
        ctk.CTkButton(
            search_frame,
            text="⟳ รีเฟรช",
            width=100,
            height=38,
            command=self.refresh_data,
            font=ctk.CTkFont(family="Kanit", size=13),
            corner_radius=RADIUS_INPUT,
            fg_color="#F8FAFC",
            border_width=1,
            border_color="#E2E8F0",
            text_color=TEXT_H3,
            hover_color="#E2E8F0"
        ).pack(side="right")

        # === ตาราง (card style) ===
        table_card = ctk.CTkFrame(
            self.content_frame, corner_radius=RADIUS_CARD,
            fg_color="#FFFFFF",
            border_width=1, border_color="#E2E8F0"
        )
        table_card.pack(fill="x", padx=L, pady=(0, M))

        columns = ("student_id", "title", "first_name", "last_name",
                    "class_room", "class_year", "parent_phone")
        self.tree = ttk.Treeview(table_card, columns=columns, show="headings", height=12)

        self.tree.heading("student_id", text="รหัสนักเรียน")
        self.tree.heading("title", text="คำนำหน้า")
        self.tree.heading("first_name", text="ชื่อ")
        self.tree.heading("last_name", text="นามสกุล")
        self.tree.heading("class_room", text="ห้อง")
        self.tree.heading("class_year", text="ปีการศึกษา")
        self.tree.heading("parent_phone", text="เบอร์ผู้ปกครอง")

        self.tree.column("student_id", width=110, anchor="center")
        self.tree.column("title", width=90, anchor="center")
        self.tree.column("first_name", width=160)
        self.tree.column("last_name", width=160)
        self.tree.column("class_room", width=80, anchor="center")
        self.tree.column("class_year", width=100, anchor="center")
        self.tree.column("parent_phone", width=130, anchor="center")

        scrollbar = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=1)
        scrollbar.pack(side="right", fill="y", pady=1, padx=(0, 1))

        self.tree.bind("<Double-1>", lambda e: self.edit_student())

        # Empty state label (จะแสดง/ซ่อนตามข้อมูล)
        self.empty_label = ctk.CTkLabel(
            table_card,
            text="📋 ยังไม่มีข้อมูลนักเรียน กด + เพื่อเพิ่ม",
            font=ctk.CTkFont(family="Kanit", size=14),
            text_color=TEXT_CAPTION
        )

        # === ปุ่มด้านล่าง ===
        button_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        button_frame.pack(fill="x", padx=L, pady=(0, L))

        # ปุ่มด้านซ้าย
        left_buttons = ctk.CTkFrame(button_frame, fg_color="transparent")
        left_buttons.pack(side="left")

        # ปุ่ม PRIMARY: เพิ่มนักเรียน (ปุ่มหลักเดียวในส่วนนี้)
        ctk.CTkButton(
            left_buttons,
            text="+ เพิ่มนักเรียน",
            command=self.add_student,
            font=ctk.CTkFont(family="Kanit", size=14, weight="500"),
            width=150,
            height=44,
            corner_radius=RADIUS_INPUT,
            fg_color=PRIMARY,
            hover_color="#2563EB"
        ).pack(side="left", padx=(0, S))

        # ปุ่ม SECONDARY: แก้ไข
        ctk.CTkButton(
            left_buttons,
            text="✏️ แก้ไข",
            command=self.edit_student,
            font=ctk.CTkFont(family="Kanit", size=13),
            width=110,
            height=40,
            corner_radius=RADIUS_INPUT,
            fg_color="#FFFFFF",
            border_width=1,
            border_color=PRIMARY,
            text_color=PRIMARY,
            hover_color="#E0F2FE"
        ).pack(side="left", padx=(0, S))

        # ปุ่ม SECONDARY: ลบ
        ctk.CTkButton(
            left_buttons,
            text="🗑️ ลบ",
            command=self.delete_student,
            font=ctk.CTkFont(family="Kanit", size=13),
            width=90,
            height=40,
            corner_radius=RADIUS_INPUT,
            fg_color="#FFFFFF",
            border_width=1,
            border_color=DANGER,
            text_color=DANGER,
            hover_color="#FEE2E2"
        ).pack(side="left")

        # ปุ่มด้านขวา (SECONDARY style ทั้งหมด)
        right_buttons = ctk.CTkFrame(button_frame, fg_color="transparent")
        right_buttons.pack(side="right")

        for text, cmd, icon_name in [("📥 Import", self.import_excel, "file-import"),
                                      ("Export Excel", self.export_excel, "file-export"),
                                      ("Export PDF", self.export_pdf, "file-pdf")]:
            ctk.CTkButton(
                right_buttons,
                text=text,
                command=cmd,
                font=ctk.CTkFont(family="TH Sarabun New", size=14),
                width=120,
                height=40,
                corner_radius=RADIUS_BUTTON,
                fg_color="transparent",
                border_width=1,
                border_color=NEUTRAL,
                text_color=NEUTRAL,
                hover_color="#F3F4F6",
                image=IconManager.get(icon_name, 14, color=NEUTRAL, dark_color="#9CA3AF"),
                compound="left"
            ).pack(side="left", padx=(S, 0))

    def setup_table_style(self):
        """ตั้งค่าตาราง: striped rows (#F9FAFB), hover (#EFF6FF), header (#F9FAFB)"""

        style = ttk.Style()
        style.theme_use("default")

        style.configure("Treeview",
                        background="#FFFFFF",
                        foreground=TEXT_BODY,
                        rowheight=40,
                        fieldbackground="#FFFFFF",
                        borderwidth=0,
                        font=("TH Sarabun New", 14))

        style.configure("Treeview.Heading",
                        background=PRIMARY,
                        foreground="#FFFFFF",
                        font=("TH Sarabun New", 14, "bold"),
                        relief="flat",
                        borderwidth=0,
                        padding=(0, 8))

        style.map("Treeview",
                  background=[("selected", "#DBEAFE")],
                  foreground=[("selected", "#1E40AF")])

        style.map("Treeview.Heading",
                  background=[("active", "#1D4ED8")])

        self.tree.tag_configure("oddrow", background="#F8FAFC")
        self.tree.tag_configure("evenrow", background="#FFFFFF")

    def load_students(self):
        """โหลดข้อมูลนักเรียนลงตาราง พร้อม empty state"""

        for item in self.tree.get_children():
            self.tree.delete(item)

        class_room = None if self.class_var.get() == "ทั้งหมด" else self.class_var.get()
        class_year = None if self.year_var.get() == "ทั้งหมด" else self.year_var.get()

        self.students_data = self.db.get_all_students(class_room=class_room, class_year=class_year)

        if not self.students_data:
            # แสดง Empty State
            self.empty_label.place(relx=0.5, rely=0.5, anchor="center")
            return
        else:
            self.empty_label.place_forget()

        # แสดงในตาราง Striped Rows
        for idx, student in enumerate(self.students_data):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            self.tree.insert("", "end", values=(
                student['student_id'],
                student['title'],
                student['first_name'],
                student['last_name'],
                student['class_room'],
                student['class_year'],
                student['parent_phone'] or "-"
            ), tags=(tag,))

        self.update_status(f"โหลดข้อมูลนักเรียน {len(self.students_data)} คน", "success")

    def search_students(self):
        """ค้นหานักเรียนแบบ realtime"""

        keyword = self.search_var.get().strip()
        if not keyword:
            self.load_students()
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        results = self.db.search_students(keyword)

        if not results:
            self.empty_label.configure(text="ไม่พบข้อมูลที่ค้นหา")
            self.empty_label.place(relx=0.5, rely=0.5, anchor="center")
            return
        else:
            self.empty_label.place_forget()

        for idx, student in enumerate(results):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            self.tree.insert("", "end", values=(
                student['student_id'],
                student['title'],
                student['first_name'],
                student['last_name'],
                student['class_room'],
                student['class_year'],
                student['parent_phone'] or "-"
            ), tags=(tag,))

        self.update_status(f"พบ {len(results)} รายการ", "info")

    def refresh_data(self):
        """รีเฟรชข้อมูล"""
        self.load_students()
        self.update_status("รีเฟรชข้อมูลเรียบร้อย", "success")

    def add_student(self):
        """เปิดฟอร์มเพิ่มนักเรียน — เติมห้องอัตโนมัติถ้าเลือก filter ไว้"""
        default_room = None
        if self.class_var.get() != "ทั้งหมด":
            default_room = self.class_var.get()
        StudentForm(self.parent, self.db, self.load_students, self.update_status,
                    default_class_room=default_room)

    def _get_selected_student_id(self):
        """ดึง student_id จากแถวที่เลือก (ใช้ index เพื่อหลีกเลี่ยง Treeview ตัดเลข 0 นำหน้า)"""
        selected = self.tree.selection()
        if not selected:
            return None
        idx = self.tree.index(selected[0])
        if idx < len(self.students_data):
            return self.students_data[idx]['student_id']
        return None

    def edit_student(self):
        """เปิดฟอร์มแก้ไขนักเรียน"""
        student_id = self._get_selected_student_id()
        if not student_id:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกนักเรียนที่ต้องการแก้ไข")
            return

        student = self.db.get_student_by_id(student_id)
        if student:
            StudentForm(self.parent, self.db, self.load_students, self.update_status, student)

    def delete_student(self):
        """ลบนักเรียน - แสดง confirmation modal พร้อมชื่อรายการ"""
        student_id = self._get_selected_student_id()
        if not student_id:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกนักเรียนที่ต้องการลบ")
            return
        student = self.db.get_student_by_id(student_id)
        name = f"{student['first_name']} {student['last_name']}" if student else student_id

        # Confirmation dialog พร้อมชื่อรายการ
        confirm = messagebox.askyesno(
            "ยืนยันการลบ",
            f"ต้องการลบนักเรียน \"{name}\" หรือไม่?\n(ข้อมูลจะถูกซ่อน ไม่ถูกลบถาวร)"
        )

        if confirm:
            if self.db.delete_student(student_id):
                self.load_students()
                self.update_status(f"ลบนักเรียน {name} เรียบร้อย", "success")
            else:
                self.update_status("ไม่สามารถลบนักเรียนได้", "error")

    def import_excel(self):
        """นำเข้าข้อมูลจาก Excel"""

        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            success_count = 0
            error_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                student_data = {
                    'student_id': str(row[0]),
                    'title': row[1] or "เด็กชาย",
                    'first_name': row[2] or "",
                    'last_name': row[3] or "",
                    'class_room': row[4] or "",
                    'class_year': str(datetime.now().year + 543),
                    'birth_date': row[5] or "",
                    'parent_name': row[6] or "",
                    'parent_phone': row[7] or "",
                    'photo_path': None
                }
                if self.db.add_student(student_data):
                    success_count += 1
                else:
                    error_count += 1

            self.load_students()
            self.update_status(f"Import สำเร็จ {success_count} คน", "success")
            messagebox.showinfo("สำเร็จ", f"นำเข้าข้อมูลเรียบร้อย\nสำเร็จ: {success_count}\nผิดพลาด: {error_count}")

        except Exception as e:
            self.update_status("ไม่สามารถนำเข้าข้อมูลได้", "error")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถนำเข้าข้อมูลได้\n{str(e)}")

    def export_excel(self):
        """ส่งออกข้อมูลเป็น Excel"""

        if not self.students_data:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลให้ส่งออก")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"รายชื่อนักเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "รายชื่อนักเรียน"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )

            headers = ["รหัสนักเรียน", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง",
                        "ปีการศึกษา", "วันเกิด", "ผู้ปกครอง", "เบอร์ติดต่อ"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for row_idx, student in enumerate(self.students_data, start=2):
                data = [
                    student['student_id'], student['title'],
                    student['first_name'], student['last_name'],
                    student['class_room'], student['class_year'],
                    student['birth_date'] or "-",
                    student['parent_name'] or "-",
                    student['parent_phone'] or "-"
                ]
                row_fill = PatternFill(start_color="F9FAFB", fill_type="solid") if row_idx % 2 == 0 else None
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    if row_fill:
                        cell.fill = row_fill
                    cell.alignment = Alignment(horizontal='center' if col_idx in [1, 2, 5, 6] else 'left')

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col_letter].width = 15

            wb.save(file_path)
            self.update_status("ส่งออก Excel สำเร็จ", "success")
            messagebox.showinfo("สำเร็จ", f"ส่งออกข้อมูล {len(self.students_data)} รายการเรียบร้อย")

        except Exception as e:
            self.update_status("ไม่สามารถส่งออกข้อมูลได้", "error")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถส่งออกข้อมูลได้\n{str(e)}")

    def export_pdf(self):
        """ส่งออกข้อมูลเป็น PDF"""

        if not self.students_data:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลให้ส่งออก")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"รายชื่อนักเรียน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Heading1'],
                fontName=font_name, fontSize=18, alignment=1
            )
            title = Paragraph("รายชื่อนักเรียน", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            data = [["รหัส", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "ปีการศึกษา", "เบอร์ติดต่อ"]]
            for student in self.students_data:
                data.append([
                    student['student_id'], student['title'],
                    student['first_name'], student['last_name'],
                    student['class_room'], student['class_year'],
                    student['parent_phone'] or "-"
                ])

            table = Table(data, colWidths=[3 * cm, 2 * cm, 3.5 * cm, 3.5 * cm, 2 * cm, 2.5 * cm, 3 * cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(PRIMARY)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F9FAFB")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(TABLE_BORDER)),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))

            elements.append(table)
            doc.build(elements)

            self.update_status("ส่งออก PDF สำเร็จ", "success")
            messagebox.showinfo("สำเร็จ", f"ส่งออกข้อมูล {len(self.students_data)} รายการเรียบร้อย")

        except Exception as e:
            self.update_status("ไม่สามารถส่งออก PDF ได้", "error")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถส่งออก PDF ได้\n{str(e)}")


class StudentForm(ctk.CTkToplevel):
    """ฟอร์มเพิ่ม/แก้ไขนักเรียน"""

    def __init__(self, parent, db, callback, update_status, student=None, default_class_room=None):
        super().__init__(parent)

        self.db = db
        self.callback = callback
        self.update_status = update_status
        self.student = student
        self.default_class_room = default_class_room
        self.error_labels = {}

        title_text = "แก้ไขนักเรียน" if student else "เพิ่มนักเรียน"
        self.title(title_text)
        self.geometry("480x580")
        self.resizable(False, False)
        self.center_window()
        self.transient(parent)
        self.grab_set()

        self.create_form()
        if student:
            self.fill_data()

    def center_window(self):
        """จัดหน้าต่างให้อยู่กึ่งกลาง"""
        self.update_idletasks()
        width, height = 480, 580
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def create_form(self):
        """สร้างฟอร์ม - label อยู่บน input อยู่ล่าง, 2 คอลัมน์"""

        main_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#FFFFFF")
        main_frame.pack(fill="both", expand=True)

        # Header
        header_frame = ctk.CTkFrame(main_frame, fg_color=PRIMARY, corner_radius=0, height=48)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        title_text = "แก้ไขข้อมูลนักเรียน" if self.student else "เพิ่มนักเรียนใหม่"
        ctk.CTkLabel(
            header_frame, text=title_text,
            font=ctk.CTkFont(family="TH Sarabun New", size=18, weight="bold"),
            text_color="white"
        ).pack(expand=True)

        # Scrollable form
        scroll = ctk.CTkScrollableFrame(
            main_frame, fg_color="transparent",
            scrollbar_button_color=NEUTRAL, scrollbar_button_hover_color=PRIMARY
        )
        scroll.pack(fill="both", expand=True, padx=L, pady=(M, 0))

        # แถว 1: รหัสนักเรียน | คำนำหน้า
        row1 = ctk.CTkFrame(scroll, fg_color="transparent")
        row1.pack(fill="x", pady=(0, S))
        row1.columnconfigure(0, weight=1)
        row1.columnconfigure(1, weight=1)
        self.create_field(row1, "รหัสนักเรียน", "student_id", required=True, grid_pos=(0, 0))
        self._create_title_field(row1, grid_pos=(0, 1))

        # แถว 2: ชื่อ | นามสกุล
        row2 = ctk.CTkFrame(scroll, fg_color="transparent")
        row2.pack(fill="x", pady=(0, S))
        row2.columnconfigure(0, weight=1)
        row2.columnconfigure(1, weight=1)
        self.create_field(row2, "ชื่อ", "first_name", required=True, grid_pos=(0, 0))
        self.create_field(row2, "นามสกุล", "last_name", required=True, grid_pos=(0, 1))

        # แถว 3: ห้อง | ปีการศึกษา
        row3 = ctk.CTkFrame(scroll, fg_color="transparent")
        row3.pack(fill="x", pady=(0, S))
        row3.columnconfigure(0, weight=1)
        row3.columnconfigure(1, weight=1)
        self.create_field(row3, "ห้อง", "class_room", placeholder="เช่น ป.1/1",
                         default=self.default_class_room or "", required=True, grid_pos=(0, 0))
        current_year = datetime.now().year + 543
        self.create_field(row3, "ปีการศึกษา", "class_year", default=str(current_year), grid_pos=(0, 1))

        # แถว 4: วันเกิด (เต็มแถว)
        self.create_field(scroll, "วันเกิด", "birth_date", placeholder="วว/ดด/ปปปป")

        # แถว 5: ชื่อผู้ปกครอง | เบอร์ผู้ปกครอง
        row5 = ctk.CTkFrame(scroll, fg_color="transparent")
        row5.pack(fill="x", pady=(0, S))
        row5.columnconfigure(0, weight=1)
        row5.columnconfigure(1, weight=1)
        self.create_field(row5, "ชื่อผู้ปกครอง", "parent_name", grid_pos=(0, 0))
        self.create_field(row5, "เบอร์ผู้ปกครอง", "parent_phone", placeholder="0812345678", grid_pos=(0, 1))

        # ปุ่มด้านล่าง (นอก scroll — เห็นตลอด)
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=M, padx=L)

        ctk.CTkLabel(
            button_frame, text="* = จำเป็นต้องกรอก",
            font=ctk.CTkFont(family="TH Sarabun New", size=12),
            text_color=TEXT_CAPTION
        ).pack(side="left")

        ctk.CTkButton(
            button_frame, text="ยกเลิก", command=self.destroy,
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            width=90, height=36, corner_radius=RADIUS_BUTTON,
            fg_color="transparent", border_width=1,
            border_color=NEUTRAL, text_color=NEUTRAL, hover_color="#F3F4F6"
        ).pack(side="right", padx=(S, 0))

        ctk.CTkButton(
            button_frame, text="บันทึก", command=self.save,
            font=ctk.CTkFont(family="TH Sarabun New", size=14, weight="bold"),
            width=90, height=36, corner_radius=RADIUS_BUTTON,
            fg_color=PRIMARY, hover_color="#1D4ED8",
            image=IconManager.get_white("floppy-disk", 14), compound="left"
        ).pack(side="right")

    def _create_title_field(self, parent, grid_pos=None):
        """สร้าง field คำนำหน้า (dropdown)"""
        container = ctk.CTkFrame(parent, fg_color="transparent")
        if grid_pos:
            container.grid(row=grid_pos[0], column=grid_pos[1], sticky="nsew",
                           padx=(S if grid_pos[1] > 0 else 0, 0))
        else:
            container.pack(fill="x", pady=(0, S))

        ctk.CTkLabel(
            container, text="คำนำหน้า *",
            font=ctk.CTkFont(family="TH Sarabun New", size=13, weight="bold"),
            text_color=TEXT_H3, anchor="w"
        ).pack(fill="x")

        self.title_var = ctk.StringVar(value="เด็กชาย")
        ctk.CTkOptionMenu(
            container, variable=self.title_var,
            values=["เด็กชาย", "เด็กหญิง", "นาย", "นางสาว", "นาง"],
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            height=36, corner_radius=20,
            fg_color="#EFF6FF", button_color="#EFF6FF",
            button_hover_color="#DBEAFE", text_color="#1E40AF",
            dropdown_fg_color="#F0F4FF", dropdown_hover_color="#DBEAFE",
            dropdown_text_color="#1E40AF",
            dropdown_font=ctk.CTkFont(family="TH Sarabun New", size=16)
        ).pack(fill="x", pady=(XS, 0))

    def create_field(self, parent, label, field_name, placeholder="", default="", required=False, grid_pos=None):
        """สร้าง field — label อยู่บน, input อยู่ล่าง"""
        container = ctk.CTkFrame(parent, fg_color="transparent")
        if grid_pos:
            container.grid(row=grid_pos[0], column=grid_pos[1], sticky="nsew",
                           padx=(S if grid_pos[1] > 0 else 0, 0))
        else:
            container.pack(fill="x", pady=(0, S))

        label_text = f"{label} *" if required else label
        ctk.CTkLabel(
            container, text=label_text,
            font=ctk.CTkFont(family="TH Sarabun New", size=13, weight="bold"),
            text_color=TEXT_H3, anchor="w"
        ).pack(fill="x")

        var = ctk.StringVar(value=default)
        setattr(self, f"{field_name}_var", var)

        entry = ctk.CTkEntry(
            container, textvariable=var, placeholder_text=placeholder,
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            height=34, corner_radius=RADIUS_BUTTON,
            border_width=1, border_color=INPUT_BORDER
        )
        entry.pack(fill="x", pady=(XS, 0))
        setattr(self, f"{field_name}_entry", entry)

        error_label = ctk.CTkLabel(
            container, text="",
            font=ctk.CTkFont(family="TH Sarabun New", size=11),
            text_color=DANGER, anchor="w", height=14
        )
        error_label.pack(fill="x", pady=0)
        self.error_labels[field_name] = error_label

    def show_field_error(self, field_name, message):
        """แสดง error ใต้ field สีแดง #DC2626 ขนาด 12px"""
        if field_name in self.error_labels:
            self.error_labels[field_name].configure(text=message)
        # เปลี่ยน border ของ entry เป็นสีแดง
        entry = getattr(self, f"{field_name}_entry", None)
        if entry:
            entry.configure(border_color=DANGER)

    def clear_field_errors(self):
        """ล้าง error ทั้งหมด"""
        for field_name, label in self.error_labels.items():
            label.configure(text="")
            entry = getattr(self, f"{field_name}_entry", None)
            if entry:
                entry.configure(border_color=INPUT_BORDER)

    def fill_data(self):
        """ใส่ข้อมูลเดิมในฟอร์ม (โหมดแก้ไข)"""
        self.student_id_var.set(self.student['student_id'])
        self.title_var.set(self.student['title'])
        self.first_name_var.set(self.student['first_name'])
        self.last_name_var.set(self.student['last_name'])
        self.class_room_var.set(self.student['class_room'])
        self.class_year_var.set(self.student['class_year'])
        self.birth_date_var.set(self.student['birth_date'] or "")
        self.parent_name_var.set(self.student['parent_name'] or "")
        self.parent_phone_var.set(self.student['parent_phone'] or "")

    def save(self):
        """บันทึกข้อมูล พร้อม validation แสดง error ใต้ field"""

        self.clear_field_errors()
        has_error = False

        # Validation
        if not self.student_id_var.get().strip():
            self.show_field_error("student_id", "กรุณากรอกรหัสนักเรียน")
            has_error = True

        if not self.first_name_var.get().strip():
            self.show_field_error("first_name", "กรุณากรอกชื่อ")
            has_error = True

        if not self.last_name_var.get().strip():
            self.show_field_error("last_name", "กรุณากรอกนามสกุล")
            has_error = True

        if not self.class_room_var.get().strip():
            self.show_field_error("class_room", "กรุณากรอกห้อง")
            has_error = True

        if has_error:
            return

        student_data = {
            'student_id': self.student_id_var.get().strip(),
            'title': self.title_var.get(),
            'first_name': self.first_name_var.get().strip(),
            'last_name': self.last_name_var.get().strip(),
            'class_room': self.class_room_var.get().strip(),
            'class_year': self.class_year_var.get().strip(),
            'birth_date': self.birth_date_var.get().strip() or None,
            'parent_name': self.parent_name_var.get().strip() or None,
            'parent_phone': self.parent_phone_var.get().strip() or None,
            'photo_path': None
        }

        if self.student:
            success = self.db.update_student(self.student['student_id'], student_data)
            message = "แก้ไขข้อมูลนักเรียนเรียบร้อย"
        else:
            success = self.db.add_student(student_data)
            message = "เพิ่มนักเรียนเรียบร้อย"

        if success:
            self.update_status(message, "success")
            self.callback()
            self.destroy()
        else:
            self.update_status("ไม่สามารถบันทึกข้อมูลได้", "error")
            messagebox.showerror("ผิดพลาด", "ไม่สามารถบันทึกข้อมูลได้\n(รหัสนักเรียนอาจซ้ำ)")
