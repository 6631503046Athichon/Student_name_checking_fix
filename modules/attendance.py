"""
modules/attendance.py
โมดูลเช็คชื่อ - Design System v3.0 (Teacher-Friendly Edition)
- ตารางเช็คชื่อแนวตั้ง ดูง่าย อ่านชัด
- ปุ่มสถานะมีข้อความ + สี ชัดเจน
- Quick Action: "เช็คมาทั้งหมด" ต่อห้อง ประหยัดเวลา
- สรุปจำนวน มา/ขาด/ลา/สาย แบบ Real-time
- Empty state, Toast notification
"""

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from modules.icons import IconManager
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from modules.pdf_utils import get_thai_font

# ==================== Design System v4.0 ====================
PRIMARY = "#3B82F6"
PRIMARY_LIGHT = "#E0F2FE"
PRIMARY_HOVER = "#2563EB"
SUCCESS = "#10B981"
SUCCESS_LIGHT = "#D1FAE5"
SUCCESS_HOVER = "#059669"
WARNING = "#F59E0B"
WARNING_LIGHT = "#FEF3C7"
WARNING_HOVER = "#D97706"
DANGER = "#EF4444"
DANGER_LIGHT = "#FEE2E2"
DANGER_HOVER = "#DC2626"
NEUTRAL = "#64748B"
NEUTRAL_LIGHT = "#F1F5F9"
NEUTRAL_HOVER = "#475569"

TEXT_H1 = "#111827"
TEXT_H2 = "#1F2937"
TEXT_H3 = "#374151"
TEXT_BODY = "#374151"
TEXT_CAPTION = "#6B7280"

TABLE_HEADER_BG = "#F9FAFB"
TABLE_HOVER = "#EFF6FF"
TABLE_STRIPE = "#F8FAFC"
TABLE_BORDER = "#E5E7EB"

XS, S, M, L, XL, XXL = 4, 8, 16, 24, 32, 48
RADIUS_BUTTON = 8
RADIUS_CARD = 12
RADIUS_MODAL = 16
RADIUS_PILL = 20

INPUT_BORDER = "#D1D5DB"

# สถานะ config - ใช้ทั่วทั้ง module
STATUSES = [
    {"text": "มา", "value": "มา", "color": SUCCESS, "light": SUCCESS_LIGHT, "hover": SUCCESS_HOVER, "icon": "circle-check"},
    {"text": "ขาด", "value": "ขาด", "color": DANGER, "light": DANGER_LIGHT, "hover": DANGER_HOVER, "icon": "circle-xmark"},
    {"text": "ลา", "value": "ลา", "color": NEUTRAL, "light": NEUTRAL_LIGHT, "hover": NEUTRAL_HOVER, "icon": "circle-minus"},
    {"text": "สาย", "value": "มาสาย", "color": WARNING, "light": WARNING_LIGHT, "hover": WARNING_HOVER, "icon": "clock"},
]
STATUS_MAP = {s["value"]: s for s in STATUSES}


class AttendanceModule:
    """โมดูลเช็คชื่อ - Teacher-Friendly Edition"""

    def __init__(self, parent, db, update_status_callback):
        self.parent = parent
        self.db = db
        self.update_status = update_status_callback
        self.current_date = datetime.now()
        self.students_data = []
        self.attendance_buttons = {}
        self.room_summary_labels = {}
        self.global_summary_labels = {}

        self.create_ui()

    def create_ui(self):
        """สร้าง UI หลัก"""

        self.content_frame = ctk.CTkScrollableFrame(
            self.parent,
            fg_color="transparent",
            scrollbar_button_color="#CBD5E1",
            scrollbar_button_hover_color=PRIMARY
        )
        self.content_frame.pack(fill="both", expand=True)

        self.tabview = ctk.CTkTabview(
            self.content_frame, corner_radius=RADIUS_CARD,
            fg_color="#FFFFFF", border_width=1, border_color="#E2E8F0",
            segmented_button_fg_color="#E2E8F0",
            segmented_button_selected_color=PRIMARY,
            segmented_button_unselected_color="#E2E8F0",
            segmented_button_selected_hover_color="#2563EB",
            segmented_button_unselected_hover_color="#CBD5E1",
            text_color="#FFFFFF",
            font=ctk.CTkFont(family="Kanit", size=13, weight="500")
        )
        self.tabview.pack(fill="both", expand=True, padx=L, pady=L)

        self.tabview.add("📝 เช็คชื่อรายวัน")
        self.create_daily_tab()

        self.tabview.add("📊 รายงานขาดเรียน")
        self.create_absent_report_tab()

    # ==================== TAB เช็คชื่อรายวัน ====================

    def create_daily_tab(self):
        """Tab เช็คชื่อรายวัน - ตาราง + Quick Actions"""

        tab = self.tabview.tab("เช็คชื่อรายวัน")

        # === แถบควบคุมด้านบน ===
        control_card = ctk.CTkFrame(
            tab, fg_color="#FFFFFF",
            corner_radius=RADIUS_CARD, border_width=1, border_color=TABLE_BORDER
        )
        control_card.pack(fill="x", padx=M, pady=(M, S))

        # แถว 1: วันที่ + ห้อง + ปุ่มโหลด
        row1 = ctk.CTkFrame(control_card, fg_color="transparent")
        row1.pack(fill="x", padx=L, pady=(L, S))

        # วันที่
        ctk.CTkLabel(
            row1, text="📅 วันที่:",
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, S))

        self.date_var = ctk.StringVar(value=self.current_date.strftime("%Y-%m-%d"))
        date_entry = ctk.CTkEntry(
            row1, textvariable=self.date_var,
            width=150, height=40,
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            corner_radius=RADIUS_BUTTON, border_width=1, border_color=INPUT_BORDER,
            placeholder_text="YYYY-MM-DD"
        )
        date_entry.pack(side="left", padx=(0, L))

        # เลือกห้อง
        ctk.CTkLabel(
            row1, text="🏫 ห้อง:",
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, S))

        self.daily_class_var = ctk.StringVar(value="ทั้งหมด")
        class_options = ["ทั้งหมด"] + self.db.get_class_rooms()

        ctk.CTkOptionMenu(
            row1,
            variable=self.daily_class_var,
            values=class_options,
            command=lambda x: self.load_daily_attendance(),
            width=160, height=40,
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            corner_radius=RADIUS_PILL,
            fg_color=PRIMARY_LIGHT, button_color=PRIMARY_LIGHT,
            button_hover_color="#DBEAFE", text_color="#1E40AF",
            dropdown_fg_color="#F0F4FF", dropdown_hover_color="#DBEAFE",
            dropdown_text_color="#1E40AF",
            dropdown_font=ctk.CTkFont(family="TH Sarabun New", size=16)
        ).pack(side="left", padx=(0, L))

        # ปุ่มโหลดข้อมูล
        ctk.CTkButton(
            row1, text="  โหลดข้อมูล",
            command=self.load_daily_attendance,
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            width=140, height=40,
            corner_radius=RADIUS_BUTTON,
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER,
            image=IconManager.get_white("rotate", 14), compound="left"
        ).pack(side="left")

        # ปุ่มบันทึก (ขวาสุด - เด่นชัด)
        save_btn = ctk.CTkButton(
            row1, text="  💾 บันทึกทั้งหมด",
            command=self.save_all_attendance,
            font=ctk.CTkFont(family="TH Sarabun New", size=18, weight="bold"),
            width=180, height=44,
            corner_radius=RADIUS_BUTTON,
            fg_color=SUCCESS, hover_color=SUCCESS_HOVER,
        )
        save_btn.pack(side="right")

        # แถว 2: สรุปจำนวน Real-time
        self.summary_frame = ctk.CTkFrame(control_card, fg_color=TABLE_STRIPE, corner_radius=RADIUS_BUTTON)
        self.summary_frame.pack(fill="x", padx=L, pady=(S, L))

        summary_inner = ctk.CTkFrame(self.summary_frame, fg_color="transparent")
        summary_inner.pack(padx=M, pady=S)

        ctk.CTkLabel(
            summary_inner, text="📊 สรุป:",
            font=ctk.CTkFont(family="TH Sarabun New", size=15, weight="bold"),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, M))

        # จำนวนรวม
        self.global_summary_labels["total"] = ctk.CTkLabel(
            summary_inner, text="ทั้งหมด 0 คน",
            font=ctk.CTkFont(family="TH Sarabun New", size=15, weight="bold"),
            text_color=TEXT_H2
        )
        self.global_summary_labels["total"].pack(side="left", padx=(0, L))

        # สรุปแต่ละสถานะ
        for st in STATUSES:
            pill = ctk.CTkFrame(summary_inner, fg_color=st["light"], corner_radius=RADIUS_PILL)
            pill.pack(side="left", padx=(0, S))

            lbl = ctk.CTkLabel(
                pill, text=f"  {st['text']}: 0  ",
                font=ctk.CTkFont(family="TH Sarabun New", size=14, weight="bold"),
                text_color=st["color"]
            )
            lbl.pack(padx=S, pady=XS)
            self.global_summary_labels[st["value"]] = lbl

        # ยังไม่เลือก
        pill_none = ctk.CTkFrame(summary_inner, fg_color="#F3F4F6", corner_radius=RADIUS_PILL)
        pill_none.pack(side="left", padx=(0, S))
        self.global_summary_labels["none"] = ctk.CTkLabel(
            pill_none, text="  ยังไม่เลือก: 0  ",
            font=ctk.CTkFont(family="TH Sarabun New", size=14, weight="bold"),
            text_color="#9CA3AF"
        )
        self.global_summary_labels["none"].pack(padx=S, pady=XS)

        # === พื้นที่ตาราง ===
        self.cards_container = ctk.CTkScrollableFrame(
            tab, fg_color="#F8FAFC",
            corner_radius=RADIUS_CARD, border_width=1, border_color=TABLE_BORDER,
            scrollbar_button_color="#D1D5DB", scrollbar_button_hover_color=PRIMARY
        )
        self.cards_container.pack(fill="both", expand=True, padx=M, pady=(0, M))

        self.load_daily_attendance()

    # ==================== โหลดข้อมูลเช็คชื่อ ====================

    def load_daily_attendance(self):
        """โหลดข้อมูลเช็คชื่อ - ตารางแยกห้อง พร้อม Quick Actions"""

        for widget in self.cards_container.winfo_children():
            widget.destroy()

        self.attendance_buttons = {}
        self.room_summary_labels = {}

        date = self.date_var.get()
        class_room = None if self.daily_class_var.get() == "ทั้งหมด" else self.daily_class_var.get()

        students = self.db.get_all_students(class_room=class_room)

        if not students:
            self._show_empty_state()
            self._update_global_summary()
            return

        attendance_records = self.db.get_attendance_by_date(date, class_room)
        attendance_dict = {rec['student_id']: rec['status'] for rec in attendance_records}

        # จัดกลุ่มตามห้อง
        rooms = {}
        for student in students:
            room = student['class_room']
            if room not in rooms:
                rooms[room] = []
            rooms[room].append(student)

        # สร้างตารางแต่ละห้อง
        for room_idx, room_name in enumerate(sorted(rooms.keys())):
            room_students = rooms[room_name]
            self._create_room_section(room_name, room_students, attendance_dict, room_idx)

        # อัปเดตสรุป
        self._update_global_summary()
        self.update_status(f"โหลดข้อมูล {len(students)} คน ({len(rooms)} ห้อง)", "success")

    def _show_empty_state(self):
        """แสดง Empty State"""
        empty_frame = ctk.CTkFrame(self.cards_container, fg_color="transparent")
        empty_frame.pack(fill="both", expand=True, pady=XXL)

        ctk.CTkLabel(
            empty_frame, text="📋",
            font=ctk.CTkFont(size=48)
        ).pack(pady=(XL, M))

        ctk.CTkLabel(
            empty_frame,
            text="ยังไม่มีข้อมูลนักเรียน",
            font=ctk.CTkFont(family="TH Sarabun New", size=20, weight="bold"),
            text_color=TEXT_H3
        ).pack()

        ctk.CTkLabel(
            empty_frame,
            text="เลือกห้องเรียนและกด 'โหลดข้อมูล' เพื่อเริ่มเช็คชื่อ",
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            text_color="#9CA3AF"
        ).pack(pady=(S, 0))

    # ==================== ส่วนห้อง ====================

    def _create_room_section(self, room_name, students, attendance_dict, room_idx):
        """สร้างส่วนของแต่ละห้อง: Header + Quick Actions + ตาราง + สรุป"""

        # Container ของห้อง
        room_container = ctk.CTkFrame(
            self.cards_container, fg_color="#FFFFFF",
            corner_radius=RADIUS_CARD, border_width=1, border_color=TABLE_BORDER
        )
        room_container.pack(fill="x", padx=S, pady=(M if room_idx == 0 else L, 0))

        # === Header ห้อง + Quick Actions ===
        header_frame = ctk.CTkFrame(room_container, fg_color=PRIMARY_LIGHT, corner_radius=0)
        header_frame.pack(fill="x")

        header_inner = ctk.CTkFrame(header_frame, fg_color="transparent")
        header_inner.pack(fill="x", padx=L, pady=M)

        # ชื่อห้อง
        ctk.CTkFrame(
            header_inner, fg_color=PRIMARY, width=5, height=28, corner_radius=2
        ).pack(side="left", padx=(0, S))

        ctk.CTkLabel(
            header_inner,
            text=f"ห้อง {room_name}",
            font=ctk.CTkFont(family="TH Sarabun New", size=20, weight="bold"),
            text_color=TEXT_H1
        ).pack(side="left")

        ctk.CTkLabel(
            header_inner,
            text=f"  ({len(students)} คน)",
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            text_color=TEXT_CAPTION
        ).pack(side="left")

        # Quick Action Buttons (ขวา)
        quick_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        quick_frame.pack(side="right")

        ctk.CTkButton(
            quick_frame, text="  ✅ เช็คมาทั้งหมด",
            command=lambda rn=room_name: self._mark_all_room(rn, "มา"),
            font=ctk.CTkFont(family="TH Sarabun New", size=14, weight="bold"),
            width=150, height=34,
            corner_radius=RADIUS_PILL,
            fg_color=SUCCESS, hover_color=SUCCESS_HOVER,
            text_color="#FFFFFF"
        ).pack(side="left", padx=(0, S))

        ctk.CTkButton(
            quick_frame, text="  🔄 ล้างทั้งหมด",
            command=lambda rn=room_name: self._clear_all_room(rn),
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            width=120, height=34,
            corner_radius=RADIUS_PILL,
            fg_color="transparent", hover_color="#FEE2E2",
            text_color=DANGER, border_width=1, border_color=DANGER
        ).pack(side="left")

        # === ตาราง Header ===
        table_header = ctk.CTkFrame(room_container, fg_color="#1E3A5F", corner_radius=0, height=44)
        table_header.pack(fill="x")
        table_header.pack_propagate(False)

        # ใช้ grid layout ให้คอลัมน์มีสัดส่วนเหมาะสม
        header_cols = [
            # (text, col_idx, weight, minsize)
            ("ลำดับ", 0, 0, 50),
            ("รหัสนักเรียน", 1, 1, 100),
            ("ชื่อ-นามสกุล", 2, 3, 180),
            ("ห้อง", 3, 0, 60),
            ("มา", 4, 1, 70),
            ("ขาด", 5, 1, 70),
            ("ลา", 6, 1, 70),
            ("สาย", 7, 1, 70),
        ]

        for text, col, weight, minsize in header_cols:
            table_header.grid_columnconfigure(col, weight=weight, minsize=minsize)
            ctk.CTkLabel(
                table_header, text=text,
                font=ctk.CTkFont(family="TH Sarabun New", size=15, weight="bold"),
                text_color="#FFFFFF", anchor="center"
            ).grid(row=0, column=col, sticky="ew", padx=1, pady=8)

        # === แถวนักเรียน ===
        for idx, student in enumerate(students):
            self._create_student_row(room_container, student, attendance_dict, idx, room_name)

        # === สรุปท้ายห้อง ===
        summary_bar = ctk.CTkFrame(room_container, fg_color="#F0F4FF", corner_radius=0)
        summary_bar.pack(fill="x")

        summary_inner = ctk.CTkFrame(summary_bar, fg_color="transparent")
        summary_inner.pack(padx=L, pady=S)

        room_labels = {}
        ctk.CTkLabel(
            summary_inner, text=f"สรุปห้อง {room_name}:",
            font=ctk.CTkFont(family="TH Sarabun New", size=14, weight="bold"),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, M))

        for st in STATUSES:
            pill = ctk.CTkFrame(summary_inner, fg_color=st["light"], corner_radius=RADIUS_PILL)
            pill.pack(side="left", padx=2)
            lbl = ctk.CTkLabel(
                pill, text=f" {st['text']}: 0 ",
                font=ctk.CTkFont(family="TH Sarabun New", size=13, weight="bold"),
                text_color=st["color"]
            )
            lbl.pack(padx=XS, pady=2)
            room_labels[st["value"]] = lbl

        self.room_summary_labels[room_name] = room_labels
        self._update_room_summary(room_name)

    def _create_student_row(self, parent, student, attendance_dict, idx, room_name):
        """สร้างแถวนักเรียน 1 คน ในตาราง (grid layout)"""

        bg_color = "#FFFFFF" if idx % 2 == 0 else TABLE_STRIPE

        row_frame = ctk.CTkFrame(parent, fg_color=bg_color, corner_radius=0, height=48)
        row_frame.pack(fill="x")
        row_frame.pack_propagate(False)

        # ใช้ grid layout เดียวกันกับ header
        col_config = [
            (0, 0, 50),   # ลำดับ
            (1, 1, 100),  # รหัส
            (2, 3, 180),  # ชื่อ
            (3, 0, 60),   # ห้อง
            (4, 1, 70),   # มา
            (5, 1, 70),   # ขาด
            (6, 1, 70),   # ลา
            (7, 1, 70),   # สาย
        ]
        for col, weight, minsize in col_config:
            row_frame.grid_columnconfigure(col, weight=weight, minsize=minsize)

        # ลำดับ
        ctk.CTkLabel(
            row_frame, text=str(idx + 1),
            font=ctk.CTkFont(family="TH Sarabun New", size=15),
            text_color=TEXT_CAPTION, anchor="center"
        ).grid(row=0, column=0, sticky="ew", padx=1, pady=6)

        # รหัส
        ctk.CTkLabel(
            row_frame, text=str(student['student_id']),
            font=ctk.CTkFont(family="TH Sarabun New", size=15),
            text_color=TEXT_BODY, anchor="center"
        ).grid(row=0, column=1, sticky="ew", padx=1, pady=6)

        # ชื่อ-นามสกุล
        name = f"{student['title']}{student['first_name']} {student['last_name']}"
        ctk.CTkLabel(
            row_frame, text=name,
            font=ctk.CTkFont(family="TH Sarabun New", size=15, weight="bold"),
            text_color=TEXT_H2, anchor="w"
        ).grid(row=0, column=2, sticky="ew", padx=(S, 1), pady=6)

        # ห้อง
        ctk.CTkLabel(
            row_frame, text=student['class_room'],
            font=ctk.CTkFont(family="TH Sarabun New", size=14),
            text_color=TEXT_BODY, anchor="center"
        ).grid(row=0, column=3, sticky="ew", padx=1, pady=6)

        # ปุ่มสถานะ 4 ปุ่ม
        current_status = attendance_dict.get(student['student_id'], None)
        student_buttons = {}

        for s_idx, st in enumerate(STATUSES):
            is_selected = (current_status == st["value"])

            btn = ctk.CTkButton(
                row_frame, text=st["text"],
                command=lambda sv=st["value"], sid=student['student_id'], rn=room_name: self._on_select_status(sid, sv, rn),
                font=ctk.CTkFont(family="TH Sarabun New", size=14, weight="bold"),
                height=34,
                corner_radius=RADIUS_PILL,
                fg_color=st["color"] if is_selected else "transparent",
                text_color="#FFFFFF" if is_selected else st["color"],
                border_width=2,
                border_color=st["color"],
                hover_color=st["color"],
            )
            btn.grid(row=0, column=4 + s_idx, sticky="ew", padx=3, pady=6)
            student_buttons[st["value"]] = btn

        self.attendance_buttons[student['student_id']] = {
            "buttons": student_buttons,
            "room": room_name
        }

        # เส้นแบ่งแถว
        ctk.CTkFrame(parent, fg_color=TABLE_BORDER, height=1).pack(fill="x")

    # ==================== Actions ====================

    def _on_select_status(self, student_id, status, room_name):
        """เมื่อกดเลือกสถานะ - อัปเดต UI + สรุป"""
        data = self.attendance_buttons.get(student_id, {})
        buttons = data.get("buttons", {})

        for sv, btn in buttons.items():
            st = STATUS_MAP[sv]
            if sv == status:
                btn.configure(fg_color=st["color"], text_color="#FFFFFF")
            else:
                btn.configure(fg_color="transparent", text_color=st["color"])

        # อัปเดตสรุป
        self._update_room_summary(room_name)
        self._update_global_summary()

    def _mark_all_room(self, room_name, status_value):
        """เช็คสถานะทั้งห้อง (เช่น มาทั้งหมด)"""
        for sid, data in self.attendance_buttons.items():
            if data["room"] == room_name:
                self._on_select_status(sid, status_value, room_name)

    def _clear_all_room(self, room_name):
        """ล้างสถานะทั้งห้อง"""
        for sid, data in self.attendance_buttons.items():
            if data["room"] == room_name:
                buttons = data.get("buttons", {})
                for sv, btn in buttons.items():
                    st = STATUS_MAP[sv]
                    btn.configure(fg_color="transparent", text_color=st["color"])

        self._update_room_summary(room_name)
        self._update_global_summary()

    def _get_selected_status(self, student_id):
        """หาสถานะที่เลือกอยู่ของนักเรียน"""
        data = self.attendance_buttons.get(student_id, {})
        buttons = data.get("buttons", {})
        for sv, btn in buttons.items():
            st = STATUS_MAP[sv]
            try:
                if btn.cget("fg_color") == st["color"]:
                    return sv
            except Exception:
                pass
        return None

    # ==================== สรุป ====================

    def _count_statuses(self, room_name=None):
        """นับจำนวนแต่ละสถานะ"""
        counts = {sv["value"]: 0 for sv in STATUSES}
        counts["none"] = 0
        total = 0

        for sid, data in self.attendance_buttons.items():
            if room_name and data["room"] != room_name:
                continue
            total += 1
            selected = self._get_selected_status(sid)
            if selected:
                counts[selected] += 1
            else:
                counts["none"] += 1

        return counts, total

    def _update_room_summary(self, room_name):
        """อัปเดตสรุปท้ายห้อง"""
        labels = self.room_summary_labels.get(room_name, {})
        counts, total = self._count_statuses(room_name)
        for st in STATUSES:
            lbl = labels.get(st["value"])
            if lbl:
                lbl.configure(text=f" {st['text']}: {counts[st['value']]} ")

    def _update_global_summary(self):
        """อัปเดตสรุปรวมด้านบน"""
        counts, total = self._count_statuses()

        total_lbl = self.global_summary_labels.get("total")
        if total_lbl:
            total_lbl.configure(text=f"ทั้งหมด {total} คน")

        for st in STATUSES:
            lbl = self.global_summary_labels.get(st["value"])
            if lbl:
                lbl.configure(text=f"  {st['text']}: {counts[st['value']]}  ")

        none_lbl = self.global_summary_labels.get("none")
        if none_lbl:
            none_lbl.configure(text=f"  ยังไม่เลือก: {counts['none']}  ")

    # ==================== บันทึก ====================

    def save_all_attendance(self):
        """บันทึกการเช็คชื่อทั้งหมด"""

        date = self.date_var.get()

        # ตรวจสอบรูปแบบวันที่
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("รูปแบบวันที่ผิด", "กรุณากรอกวันที่ในรูปแบบ YYYY-MM-DD\nเช่น 2026-02-24")
            return

        success_count = 0
        skip_count = 0

        for student_id, data in self.attendance_buttons.items():
            selected = self._get_selected_status(student_id)
            if selected:
                if self.db.save_attendance(student_id, date, selected):
                    success_count += 1
            else:
                skip_count += 1

        if success_count > 0:
            msg = f"✅ บันทึกสำเร็จ {success_count} คน"
            if skip_count > 0:
                msg += f" (ยังไม่ได้เลือก {skip_count} คน)"
            self.update_status(msg, "success")
        else:
            self.update_status("⚠️ กรุณาเลือกสถานะอย่างน้อย 1 คนก่อนบันทึก", "warning")

    # ==================== TAB รายงานขาดเรียน ====================

    def create_absent_report_tab(self):
        """Tab รายงานขาดเรียน"""

        tab = self.tabview.tab("รายงานขาดเรียน")

        # Control Card
        control_card = ctk.CTkFrame(
            tab, fg_color="#FFFFFF",
            corner_radius=RADIUS_CARD, border_width=1, border_color=TABLE_BORDER
        )
        control_card.pack(fill="x", padx=M, pady=(M, S))

        top_frame = ctk.CTkFrame(control_card, fg_color="transparent")
        top_frame.pack(fill="x", padx=L, pady=L)

        ctk.CTkLabel(
            top_frame, text="📋 แสดงนักเรียนที่ขาดมากกว่า:",
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            text_color=TEXT_H3
        ).pack(side="left", padx=(0, S))

        self.absent_days_var = ctk.StringVar(value="3")
        ctk.CTkEntry(
            top_frame, textvariable=self.absent_days_var,
            width=60, height=40,
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            corner_radius=RADIUS_BUTTON, border_width=1, border_color=INPUT_BORDER,
            justify="center"
        ).pack(side="left", padx=(0, S))

        ctk.CTkLabel(
            top_frame, text="วัน",
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            text_color=TEXT_BODY
        ).pack(side="left", padx=(0, L))

        # ห้อง
        ctk.CTkLabel(
            top_frame, text="🏫 ห้อง:",
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            text_color=TEXT_BODY
        ).pack(side="left", padx=(0, S))

        self.absent_class_var = ctk.StringVar(value="ทั้งหมด")
        class_options = ["ทั้งหมด"] + self.db.get_class_rooms()
        ctk.CTkOptionMenu(
            top_frame, variable=self.absent_class_var,
            values=class_options, width=140, height=40,
            font=ctk.CTkFont(family="TH Sarabun New", size=16),
            corner_radius=RADIUS_PILL,
            fg_color=PRIMARY_LIGHT, button_color=PRIMARY_LIGHT,
            button_hover_color="#DBEAFE", text_color="#1E40AF",
            dropdown_fg_color="#F0F4FF", dropdown_hover_color="#DBEAFE",
            dropdown_text_color="#1E40AF",
            dropdown_font=ctk.CTkFont(family="TH Sarabun New", size=16)
        ).pack(side="left", padx=(0, L))

        ctk.CTkButton(
            top_frame, text="  🔍 ค้นหา",
            command=self.load_absent_report,
            font=ctk.CTkFont(family="TH Sarabun New", size=16, weight="bold"),
            width=120, height=40,
            corner_radius=RADIUS_BUTTON,
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER,
        ).pack(side="left")

        # ตาราง
        table_frame = ctk.CTkFrame(
            tab, corner_radius=RADIUS_CARD,
            fg_color="#FFFFFF",
            border_width=1, border_color=TABLE_BORDER
        )
        table_frame.pack(fill="both", expand=True, padx=M, pady=(0, S))

        columns = ("student_id", "name", "class_room", "absent_days")
        self.absent_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=18)

        self.absent_tree.heading("student_id", text="รหัส")
        self.absent_tree.heading("name", text="ชื่อ-นามสกุล")
        self.absent_tree.heading("class_room", text="ห้อง")
        self.absent_tree.heading("absent_days", text="จำนวนวันขาด")

        self.absent_tree.column("student_id", width=100, anchor="center")
        self.absent_tree.column("name", width=250)
        self.absent_tree.column("class_room", width=150, anchor="center")
        self.absent_tree.column("absent_days", width=150, anchor="center")

        self.setup_table_style()

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.absent_tree.yview)
        self.absent_tree.configure(yscrollcommand=scrollbar.set)

        self.absent_tree.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=1)
        scrollbar.pack(side="right", fill="y", pady=1, padx=(0, 1))

        # ปุ่ม Export
        export_frame = ctk.CTkFrame(tab, fg_color="transparent")
        export_frame.pack(fill="x", padx=M, pady=(0, M))

        for text, cmd, icon_name in [("📥 Export Excel", self.export_attendance_excel, "file-export"),
                                      ("📄 Export PDF", self.export_attendance_pdf, "file-pdf")]:
            ctk.CTkButton(
                export_frame, text=text, command=cmd,
                font=ctk.CTkFont(family="TH Sarabun New", size=15, weight="bold"),
                width=140, height=38,
                corner_radius=RADIUS_BUTTON,
                fg_color="transparent", border_width=1,
                border_color=NEUTRAL, text_color=NEUTRAL,
                hover_color="#F3F4F6",
            ).pack(side="left", padx=(0, S))

    def setup_table_style(self):
        """ตาราง striped + hover"""
        style = ttk.Style()
        style.theme_use("default")

        style.configure("Treeview",
                        background="#FFFFFF", foreground=TEXT_BODY,
                        rowheight=42, fieldbackground="#FFFFFF",
                        borderwidth=0, font=("TH Sarabun New", 15))
        style.configure("Treeview.Heading",
                        background="#1E3A5F", foreground="#FFFFFF",
                        font=("TH Sarabun New", 15, "bold"),
                        relief="flat", borderwidth=0, padding=(0, 10))
        style.map("Treeview",
                  background=[("selected", "#DBEAFE")],
                  foreground=[("selected", "#1E40AF")])
        style.map("Treeview.Heading",
                  background=[("active", PRIMARY_HOVER)])

        self.absent_tree.tag_configure("oddrow", background=TABLE_STRIPE)
        self.absent_tree.tag_configure("evenrow", background="#FFFFFF")

    def load_absent_report(self):
        """โหลดรายงานขาดเรียน"""

        for item in self.absent_tree.get_children():
            self.absent_tree.delete(item)

        try:
            days = int(self.absent_days_var.get())
        except ValueError:
            messagebox.showwarning("คำเตือน", "กรุณากรอกจำนวนวันเป็นตัวเลข")
            return

        class_room = None if self.absent_class_var.get() == "ทั้งหมด" else self.absent_class_var.get()
        students = self.db.get_students_absent_more_than(days, class_room)

        for idx, student in enumerate(students):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            name = f"{student['title']}{student['first_name']} {student['last_name']}"
            self.absent_tree.insert("", "end", values=(
                student['student_id'], name,
                student['class_room'], student['absent_days']
            ), tags=(tag,))

        self.update_status(f"📋 พบนักเรียนขาดมากกว่า {days} วัน: {len(students)} คน", "info")

    # ==================== Export ====================

    def export_attendance_excel(self):
        """Export การเช็คชื่อเป็น Excel"""

        date_str = self.current_date.strftime("%Y-%m-%d")
        records = self.db.get_attendance_by_date(date_str)
        if not records:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลการเช็คชื่อวันนี้")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"เช็คชื่อ_{self.current_date.strftime('%Y%m%d')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "เช็คชื่อ"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            status_fills = {
                "มา": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                "ขาด": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
                "ลา": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
                "มาสาย": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            }

            headers = ["รหัสนักเรียน", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "สถานะ", "หมายเหตุ"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for row_idx, record in enumerate(records, start=2):
                data = [
                    record['student_id'],
                    record.get('title', ''),
                    record.get('first_name', ''),
                    record.get('last_name', ''),
                    record.get('class_room', ''),
                    record.get('status', ''),
                    record.get('note', '') or '',
                ]
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx in [1, 5, 6] else 'left')
                    if row_idx % 2 == 0:
                        cell.fill = stripe_fill
                    if col_idx == 6 and value in status_fills:
                        cell.fill = status_fills[value]

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col_letter].width = 15

            wb.save(file_path)
            self.update_status(f"📥 Export เช็คชื่อ {len(records)} รายการสำเร็จ", "success")
            messagebox.showinfo("สำเร็จ", f"Export ข้อมูลเช็คชื่อ {len(records)} รายการเรียบร้อย")

        except Exception as e:
            self.update_status("ไม่สามารถ Export ได้", "error")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export ได้\n{str(e)}")

    def export_attendance_pdf(self):
        """Export การเช็คชื่อเป็น PDF"""

        date_str = self.current_date.strftime("%Y-%m-%d")
        records = self.db.get_attendance_by_date(date_str)
        if not records:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลการเช็คชื่อวันนี้")
            return

        file_path = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"เช็คชื่อ_{self.current_date.strftime('%Y%m%d')}.pdf"
        )
        if not file_path:
            return

        try:
            font_name = get_thai_font()

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            thai_date = self.current_date.strftime('%d/%m/%Y')
            title = Paragraph(f"รายงานการเช็คชื่อ วันที่ {thai_date}", ParagraphStyle(
                'Title', parent=styles['Heading1'],
                fontName=font_name, fontSize=18, alignment=1
            ))
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            data = [["รหัส", "ชื่อ-สกุล", "ห้อง", "สถานะ", "หมายเหตุ"]]
            for r in records:
                name = f"{r.get('title', '')}{r.get('first_name', '')} {r.get('last_name', '')}"
                data.append([
                    r['student_id'],
                    name,
                    r.get('class_room', ''),
                    r.get('status', ''),
                    r.get('note', '') or '-',
                ])

            table = Table(data, colWidths=[3 * cm, 5 * cm, 2.5 * cm, 2.5 * cm, 4 * cm])

            style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]

            status_colors = {
                "มา": "#DCFCE7", "ขาด": "#FEE2E2",
                "ลา": "#F3F4F6", "มาสาย": "#FEF3C7",
            }
            for row_idx, r in enumerate(records, start=1):
                status = r.get('status', '')
                if status in status_colors:
                    style_cmds.append(
                        ('BACKGROUND', (3, row_idx), (3, row_idx), colors.HexColor(status_colors[status]))
                    )

            table.setStyle(TableStyle(style_cmds))

            elements.append(table)
            elements.append(Spacer(1, 0.5 * cm))

            status_count = {}
            for r in records:
                s = r.get('status', 'ไม่ระบุ')
                status_count[s] = status_count.get(s, 0) + 1
            summary_text = "สรุป: " + ", ".join(f"{k} {v} คน" for k, v in status_count.items())
            elements.append(Paragraph(summary_text, ParagraphStyle(
                'Summary', parent=styles['Normal'],
                fontName=font_name, fontSize=12, alignment=0
            )))

            doc.build(elements)
            self.update_status(f"📄 Export PDF เช็คชื่อสำเร็จ", "success")
            messagebox.showinfo("สำเร็จ", f"Export PDF เช็คชื่อ {len(records)} รายการเรียบร้อย")

        except Exception as e:
            self.update_status("ไม่สามารถ Export PDF ได้", "error")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถ Export PDF ได้\n{str(e)}")
